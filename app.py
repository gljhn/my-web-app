from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from datetime import datetime, timedelta
import json
import os
import hashlib
import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io
import secrets
import logging
from logging.handlers import RotatingFileHandler
import sys
import time
from functools import wraps
from mysql.connector import pooling


# 如果使用dateutil，需要安装：pip install python-dateutil
try:
    from dateutil import parser
    HAS_DATEUTIL = True
except ImportError:
    HAS_DATEUTIL = False
    logger.info("未安装python-dateutil，使用内置日期解析")

# 配置日志
def setup_logging():
    # 创建logs目录
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # 配置根日志记录器
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            RotatingFileHandler('logs/app.log', maxBytes=1024*1024, backupCount=5),
            logging.StreamHandler(sys.stdout)
        ]
    )

# 在导入其他模块之前设置日志
setup_logging()
logger = logging.getLogger(__name__)

# 尝试导入拼音库
try:
    from pypinyin import pinyin, Style
    HAS_PINYIN = True
    logger.info("pypinyin库加载成功")
except ImportError:
    HAS_PINYIN = False
    logger.warning("未安装pypinyin库，姓名拼音排序将使用简单排序")

# 数据库配置
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'Gn09182813',
    'database': 'web_gift_management_system',
    'charset': 'utf8mb4',
    'autocommit': True
}

# 操作类型常量
OPERATION_TYPES = {
    "ADD": "添加记录",
    "EDIT": "修改记录", 
    "DELETE": "删除记录",
    "IMPORT": "导入数据",
    "EXPORT": "导出数据",
    "LOGIN": "用户登录",
    "PASSWORD_CHANGE": "修改密码",
    "PASSWORD_RESET": "重置密码",
    "SYSTEM": "系统操作"
}

app = Flask(__name__)
app.secret_key = 'gift-management-system-secret-key-2024'
app.permanent_session_lifetime = timedelta(minutes=30)  # 会话30分钟过期

# 密码加密相关函数
def generate_salt():
    """生成随机盐值"""
    return secrets.token_hex(16)

def hash_password(password, salt):
    """使用盐值对密码进行哈希"""
    return hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000).hex()

def verify_password(stored_hash, stored_salt, provided_password):
    """验证密码"""
    return stored_hash == hash_password(provided_password, stored_salt)

def encrypt_password(password):
    """加密密码并返回哈希值和盐值"""
    salt = generate_salt()
    password_hash = hash_password(password, salt)
    return password_hash, salt

# 添加登录检查装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return jsonify({'error': '未登录或会话已过期'}), 401
        
        # 检查会话是否过期
        if 'last_activity' in session:
            last_activity = session['last_activity']
            if time.time() - last_activity > 30 * 60:  # 30分钟无操作自动登出
                session.clear()
                return jsonify({'error': '会话已过期，请重新登录'}), 401
        
        # 更新最后活动时间
        session['last_activity'] = time.time()
        return f(*args, **kwargs)
    return decorated_function

def create_connection():
    """创建数据库连接"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        logger.error(f"数据库连接失败: {str(e)}")
        return None

def safe_execute(cursor, query, params=None):
    """安全执行SQL查询"""
    try:
        cursor.execute(query, params or ())
        if cursor.with_rows:
            result = cursor.fetchall()
            return result
        return None
    except Error as e:
        raise e

def init_database():
    """初始化数据库和表"""
    connection = None
    cursor = None

    try:
        temp_config = DB_CONFIG.copy()
        temp_config.pop('database', None)
        connection = mysql.connector.connect(**temp_config)
        cursor = connection.cursor()

        safe_execute(cursor, f"CREATE DATABASE IF NOT EXISTS {DB_CONFIG['database']} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        connection.commit()

        safe_execute(cursor, f"USE {DB_CONFIG['database']}")
        connection.commit()

    # 添加更多索引
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_gift_records_date ON gift_records(date)",
            "CREATE INDEX IF NOT EXISTS idx_gift_records_type ON gift_records(record_type)",
            "CREATE INDEX IF NOT EXISTS idx_gift_records_owner ON gift_records(owner)",
            "CREATE INDEX IF NOT EXISTS idx_gift_records_name ON gift_records(name)",
            "CREATE INDEX IF NOT EXISTS idx_daily_accounts_date ON daily_accounts(account_date)",
            "CREATE INDEX IF NOT EXISTS idx_daily_accounts_type ON daily_accounts(record_type)",
            "CREATE INDEX IF NOT EXISTS idx_daily_accounts_category ON daily_accounts(category)",
            "CREATE INDEX IF NOT EXISTS idx_daily_accounts_owner ON daily_accounts(owner)",
            "CREATE INDEX IF NOT EXISTS idx_daily_accounts_type_date ON daily_accounts(record_type, account_date)",
            "CREATE INDEX IF NOT EXISTS idx_daily_accounts_owner_date ON daily_accounts(owner, account_date)",
            "CREATE INDEX IF NOT EXISTS idx_system_logs_created_at ON system_logs(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_system_logs_operation_type ON system_logs(operation_type)",
            "CREATE INDEX IF NOT EXISTS idx_system_logs_user_name ON system_logs(user_name)"
        ]
    
        for index_sql in indexes:
            try:
                safe_execute(cursor, index_sql)
                connection.commit()
                logger.info(f"成功创建索引: {index_sql.split(' ON ')[0]}")
            except Error as e:
                if "Duplicate key name" not in str(e) and "already exists" not in str(e):
                    logger.warning(f"创建索引时出现错误: {e}")

        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS gift_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                record_type ENUM('受礼记录', '随礼记录') NOT NULL DEFAULT '受礼记录',
                name VARCHAR(100) NOT NULL,
                amount DECIMAL(10,2) NOT NULL,
                occasion VARCHAR(100) NOT NULL,
                date DATE NOT NULL,
                has_returned BOOLEAN NOT NULL DEFAULT FALSE,
                return_amount DECIMAL(10,2) DEFAULT 0.00,
                return_occasion VARCHAR(100),
                return_date DATE,
                remark TEXT,
                owner VARCHAR(50) DEFAULT '郭宁',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        connection.commit()
        
        # 创建 system_config 表（确保存在）
        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS system_config (
                id INT AUTO_INCREMENT PRIMARY KEY,
                config_key VARCHAR(50) UNIQUE NOT NULL,
                config_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        connection.commit()

        try:
            safe_execute(cursor, "SELECT record_type FROM gift_records LIMIT 1")
        except Error as e:
            if "Unknown column 'record_type'" in str(e):
                safe_execute(cursor, """
                    ALTER TABLE gift_records 
                    ADD COLUMN record_type ENUM('受礼记录', '随礼记录') NOT NULL DEFAULT '受礼记录'
                """)
                connection.commit()

        try:
            safe_execute(cursor, "SELECT owner FROM gift_records LIMIT 1")
        except Error as e:
            if "Unknown column 'owner'" in str(e):
                safe_execute(cursor, """
                    ALTER TABLE gift_records 
                    ADD COLUMN owner VARCHAR(50) DEFAULT '郭宁'
                """)
                connection.commit()

        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS system_config (
                id INT AUTO_INCREMENT PRIMARY KEY,
                config_key VARCHAR(50) UNIQUE NOT NULL,
                config_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        connection.commit()

        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS system_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                operation_type VARCHAR(50) NOT NULL,
                operation_details TEXT NOT NULL,
                user_name VARCHAR(100) DEFAULT 'admin',
                record_id INT NULL,
                ip_address VARCHAR(45) DEFAULT '127.0.0.1',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_operation_type (operation_type),
                INDEX idx_created_at (created_at),
                INDEX idx_user_name (user_name)
            )
        """)
        connection.commit()
        


        # 创建用户安全信息表
        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS user_security (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password_hash VARCHAR(128) NOT NULL,
                password_salt VARCHAR(32) NOT NULL,
                security_question VARCHAR(255) NOT NULL,
                security_answer_hash VARCHAR(128) NOT NULL,
                security_answer_salt VARCHAR(32) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_username (username)
            )
        """)
        connection.commit()
        
        # ===================== 新增：记账相关表 =====================
        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS daily_accounts (
                id INT AUTO_INCREMENT PRIMARY KEY,
                record_type ENUM('支出', '收入') NOT NULL DEFAULT '支出',
                category VARCHAR(50) NOT NULL,
                subcategory VARCHAR(50),
                amount DECIMAL(10,2) NOT NULL,
                account_date DATE NOT NULL,
                description TEXT,
                payment_method VARCHAR(50) DEFAULT '现金',
                owner VARCHAR(50) DEFAULT '郭宁',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_account_date (account_date),
                INDEX idx_category (category),
                INDEX idx_record_type (record_type),
                INDEX idx_owner (owner)
            )
        """)
        connection.commit()
        
        safe_execute(cursor, """
            CREATE TABLE IF NOT EXISTS account_categories (
                id INT AUTO_INCREMENT PRIMARY KEY,
                category_type ENUM('支出', '收入') NOT NULL,
                category_name VARCHAR(50) NOT NULL,
                subcategories JSON,
                sort_order INT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_category (category_type, category_name)
            )
        """)
        connection.commit()
        
        # 使用错误处理来创建索引，如果已存在会忽略错误
        indexes = [
            "CREATE INDEX idx_gift_records_date ON gift_records(date)",
            "CREATE INDEX idx_gift_records_type ON gift_records(record_type)",
            "CREATE INDEX idx_gift_records_owner ON gift_records(owner)",
            "CREATE INDEX idx_gift_records_name ON gift_records(name)",
            "CREATE INDEX idx_daily_accounts_date ON daily_accounts(account_date)",
            "CREATE INDEX idx_daily_accounts_type ON daily_accounts(record_type)",
            "CREATE INDEX idx_daily_accounts_category ON daily_accounts(category)",
            "CREATE INDEX idx_system_logs_created_at ON system_logs(created_at)",
            "CREATE INDEX idx_system_logs_operation_type ON system_logs(operation_type)"
        ]
        
        for index_sql in indexes:
            try:
                safe_execute(cursor, index_sql)
                connection.commit()
                logger.info(f"成功创建索引: {index_sql.split(' ON ')[0]}")
            except Error as e:
                # 如果索引已存在，忽略错误；其他错误才记录
                if "Duplicate key name" not in str(e) and "already exists" not in str(e):
                    logger.warning(f"创建索引时出现错误（可能已存在）: {e}")
        
        # ===================== 修复结束 =====================
        
        # 初始化默认类别
        init_account_categories(cursor)
        
        connection.commit()

        cursor.close()
        connection.close()

        init_config()
        return True
    except Error as e:
        logger.error(f"数据库初始化失败: {str(e)}")
        return False

def get_default_username():
    """获取默认用户名"""
    connection = create_connection()
    if not connection:
        return 'admin'

    cursor = None
    try:
        cursor = connection.cursor()
        # 查找存在的用户名
        cursor.execute("SELECT username FROM user_security LIMIT 1")
        result = cursor.fetchone()
        
        if result:
            return result[0]
        else:
            return 'admin'  # 默认用户名
            
    except Error as e:
        logger.error(f"获取默认用户名错误: {e}")
        return 'admin'
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def verify_login(username, password):
    """验证用户名和密码是否正确"""
    logger.info(f"开始验证登录 - 用户名: {username}")
    
    connection = create_connection()
    if not connection:
        logger.error("数据库连接失败")
        return False

    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)
        
        # 首先尝试从user_security表查询用户信息
        logger.info(f"查询user_security表获取用户: {username}")
        cursor.execute("""
            SELECT password_hash, password_salt 
            FROM user_security 
            WHERE username = %s
        """, (username,))
        
        result = cursor.fetchone()
        
        # 如果在user_security表中找到用户，验证密码
        if result:
            logger.info(f"在user_security表中找到用户 {username}")
            logger.info(f"存储的密码哈希: {result['password_hash'][:16]}...")
            logger.info(f"存储的密码盐值: {result['password_salt'][:16]}...")
            
            is_valid = verify_password(result['password_hash'], result['password_salt'], password)
            
            if is_valid:
                logger.info(f"用户 {username} 密码验证成功")
            else:
                logger.info(f"用户 {username} 密码验证失败")
                
            return is_valid
        
        # 如果在user_security表中找不到用户，尝试从旧的system_config表验证
        logger.info(f"在user_security表中未找到用户 {username}，尝试从system_config表验证")
        cursor.execute("""
            SELECT config_value FROM system_config 
            WHERE config_key = %s
        """, (f'password_{username}',))
        
        result = cursor.fetchone()
        
        # 如果找不到对应用户名的密码，使用默认密码（向后兼容）
        if not result:
            # 检查是否有旧的单一密码配置
            cursor.execute("SELECT config_value FROM system_config WHERE config_key = 'password'")
            old_password_result = cursor.fetchone()
            
            if old_password_result and username == 'admin':
                # 向后兼容：如果用户名为admin且存在旧密码配置
                logger.info("使用system_config表中的旧密码配置")
                return password == old_password_result[0]  # 直接比较明文
            else:
                logger.info("在system_config表中也未找到用户")
                return False
        
        logger.info("使用system_config表中的密码配置")
        return password == result[0]  # 直接比较明文
        
    except Error as e:
        logger.error(f"验证登录错误: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：记账类别管理API =====================
@app.route('/api/account/categories', methods=['POST'])
@login_required
def update_account_categories():
    """更新记账类别"""
    data = request.json
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor()
        
        # 清空现有类别
        cursor.execute("DELETE FROM account_categories")
        
        # 插入新的类别数据
        categories = data.get('categories', [])
        for i, category in enumerate(categories):
            cursor.execute("""
                INSERT INTO account_categories (category_type, category_name, subcategories, sort_order)
                VALUES (%s, %s, %s, %s)
            """, (
                category['category_type'],
                category['category_name'],
                json.dumps(category['subcategories'], ensure_ascii=False),
                i
            ))
        
        connection.commit()
        
        # 记录操作日志
        log_operation("SYSTEM", "更新记账类别", user_name=session.get('username', 'admin'))
        
        return jsonify({'success': True, 'message': '类别更新成功'})
        
    except Error as e:
        logger.error(f"更新记账类别错误: {e}")
        connection.rollback()
        return jsonify({'error': '更新类别失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/account/categories/reset', methods=['POST'])
@login_required
def reset_account_categories():
    """重置记账类别为默认值"""
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor()
        
        # 清空现有类别
        cursor.execute("DELETE FROM account_categories")
        
        # 重新初始化类别
        init_account_categories(cursor)
        
        connection.commit()
        
        # 记录操作日志
        log_operation("SYSTEM", "重置记账类别为默认值", user_name=session.get('username', 'admin'))
        
        return jsonify({'success': True, 'message': '类别重置成功'})
        
    except Error as e:
        logger.error(f"重置记账类别错误: {e}")
        connection.rollback()
        return jsonify({'error': '重置类别失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 修改：记账类别初始化函数 =====================
def init_account_categories(cursor):
    """初始化记账类别"""
    # 检查是否已存在类别
    cursor.execute("SELECT COUNT(*) FROM account_categories")
    if cursor.fetchone()[0] > 0:
        return
    
    # 支出类别 - 使用您修改后的数据
    expense_categories = [
        ('食品酒水', ['早餐', '午餐', '晚餐', '粮油','调味品','水果', '零食', '烟酒']),
        ('衣服饰品', ['衣服', '裤子', '鞋子', '饰品', '化妆品']),
        ('居家物业', ['房租', '水电费', '物业费', '维修费', '日用品']),
        ('行车交通', ['公交', '地铁','铁路','共享单车', '充电桩充电','出租车', '油费', '停车费', '维修保养']),
        ('交流通讯', ['话费', '网费', '邮费']),
        ('休闲娱乐', ['电影', '旅游', '游戏', '运动', '聚会']),
        ('学习进修', ['书籍', '培训', '报名费','学费']),
        ('人情往来', ['送礼','礼品', '请客', '红包']),
        ('医疗保健', ['药品', '看病', '体检', '保健品']),
        ('金融保险', ['保险费', '手续费', '利息']),
        ('其他杂项', ['其他支出'])
    ]
    
    # 收入类别
    income_categories = [
        ('工资收入', ['工资', '奖金', '津贴']),
        ('投资收益', ['股票', '基金', '理财']),
        ('其他收入', ['兼职', '礼金', '退款']),
    ]
    
    # 插入支出类别
    for i, (category, subcategories) in enumerate(expense_categories):
        cursor.execute("""
            INSERT INTO account_categories (category_type, category_name, subcategories, sort_order)
            VALUES (%s, %s, %s, %s)
        """, ('支出', category, json.dumps(subcategories, ensure_ascii=False), i))
    
    # 插入收入类别
    for i, (category, subcategories) in enumerate(income_categories):
        cursor.execute("""
            INSERT INTO account_categories (category_type, category_name, subcategories, sort_order)
            VALUES (%s, %s, %s, %s)
        """, ('收入', category, json.dumps(subcategories, ensure_ascii=False), i))

def init_config():
    """初始化配置文件"""
    connection = create_connection()
    if not connection:
        return

    try:
        cursor = connection.cursor()
        
        # 检查是否已经存在用户
        cursor.execute("SELECT username FROM user_security WHERE username = 'admin'")
        existing_admin = cursor.fetchone()
        
        if not existing_admin:
            # 初始化默认用户（admin）
            default_username = "admin"
            default_password = "123456"
            default_security_question = "程序编写人姓氏是什么？"
            default_security_answer = "郭"
            
            # 加密密码
            password_hash, password_salt = encrypt_password(default_password)
            # 加密安全问题答案
            answer_hash, answer_salt = encrypt_password(default_security_answer)
            
            cursor.execute("""
                INSERT INTO user_security 
                (username, password_hash, password_salt, security_question, security_answer_hash, security_answer_salt)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (default_username, password_hash, password_salt, default_security_question, answer_hash, answer_salt))
            
            connection.commit()
            logger.info("初始化默认用户完成 - 密码使用哈希存储")
        
        cursor.close()
    except Error as e:
        logger.error(f"初始化配置错误: {e}")
    finally:
        if connection and connection.is_connected():
            connection.close()

def get_user_security_question(username):
    """获取用户的安全问题"""
    connection = create_connection()
    if not connection:
        return None

    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT security_question FROM user_security WHERE username = %s", (username,))
        result = cursor.fetchone()
        
        if result:
            return result[0]
        else:
            return None
            
    except Error as e:
        logger.error(f"获取安全问题错误: {e}")
        return None
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def verify_security_answer(username, answer):
    """验证安全问题答案"""
    connection = create_connection()
    if not connection:
        return False

    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT security_answer_hash, security_answer_salt 
            FROM user_security 
            WHERE username = %s
        """, (username,))
        
        result = cursor.fetchone()
        
        if not result:
            return False
        
        return verify_password(result['security_answer_hash'], result['security_answer_salt'], answer)
        
    except Error as e:
        logger.error(f"验证安全问题错误: {e}")
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def reset_user_password(username, new_password):
    """重置用户密码"""
    logger.info(f"开始重置密码 - 用户名: {username}, 新密码: '{new_password}'")
    
    connection = create_connection()
    if not connection:
        logger.error("数据库连接失败")
        return False

    cursor = None
    try:
        cursor = connection.cursor()
        
        # 生成全新的盐值
        new_salt = generate_salt()
        logger.info(f"生成的盐值: {new_salt}")
        
        # 计算新密码的哈希
        new_hash = hash_password(new_password, new_salt)
        logger.info(f"计算的新哈希: {new_hash}")
        
        # 立即验证哈希计算是否正确
        test_hash = hash_password(new_password, new_salt)
        if test_hash != new_hash:
            logger.error("❌ 哈希计算不一致！")
            return False
        
        logger.info(f"✅ 哈希计算验证通过")
        
        # 更新 user_security 表
        cursor.execute("SELECT id FROM user_security WHERE username = %s", (username,))
        user_exists = cursor.fetchone()
        
        update_success = False
        if user_exists:
            # 更新现有用户
            cursor.execute("""
                UPDATE user_security 
                SET password_hash = %s, password_salt = %s, updated_at = CURRENT_TIMESTAMP
                WHERE username = %s
            """, (new_hash, new_salt, username))
            update_success = cursor.rowcount > 0
            logger.info(f"✅ 更新 user_security 表完成，影响行数: {cursor.rowcount}")
        else:
            # 创建新用户
            default_security_question = "程序编写人姓氏是什么？"
            default_security_answer = "郭"
            
            # 加密安全问题答案
            answer_hash, answer_salt = encrypt_password(default_security_answer)
            
            cursor.execute("""
                INSERT INTO user_security 
                (username, password_hash, password_salt, security_question, security_answer_hash, security_answer_salt)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (username, new_hash, new_salt, default_security_question, answer_hash, answer_salt))
            
            update_success = cursor.rowcount > 0
            logger.info(f"✅ 创建新用户完成，影响行数: {cursor.rowcount}")
        
        # 验证更新后的数据
        cursor.execute("SELECT password_hash, password_salt FROM user_security WHERE username = %s", (username,))
        updated_data = cursor.fetchone()
        
        verification_result = False
        if updated_data:
            stored_hash, stored_salt = updated_data
            verification_hash = hash_password(new_password, stored_salt)
            verification_result = (verification_hash == stored_hash)
            
            logger.info(f"🔍 立即验证结果: {verification_result}")
            logger.info(f"  存储的哈希: {stored_hash}")
            logger.info(f"  存储的盐值: {stored_salt}")
            logger.info(f"  验证计算的哈希: {verification_hash}")
        
        if not verification_result:
            logger.error("❌ 密码重置后验证失败！")
            connection.rollback()
            return False
        else:
            logger.info("🎉 密码重置成功并验证通过！")
            connection.commit()
        
        # 同时更新 system_config 表保持兼容
        try:
            cursor.execute("""
                UPDATE system_config 
                SET config_value = %s
                WHERE config_key = %s
            """, (new_password, f'password_{username}'))
            
            if cursor.rowcount == 0:
                cursor.execute("""
                    INSERT INTO system_config (config_key, config_value)
                    VALUES (%s, %s)
                """, (f'password_{username}', new_password))
            
            logger.info(f"✅ 同时更新 system_config 表完成")
            connection.commit()
        except Exception as e:
            logger.warning(f"更新 system_config 表时出错: {e}")
            # 不因这个错误而失败
        
        # 记录操作日志
        log_operation("PASSWORD_RESET", f"通过安全问题重置密码 - 用户名: {username}", user_name=username)
        
        return True
        
    except Exception as e:
        logger.error(f"❌ 密码重置过程中出错: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def log_operation(operation_type, operation_details, record_id=None, user_name="admin", record_data=None):
    """记录系统操作日志，并自动清理一周前的旧日志"""
    connection = create_connection()
    if not connection:
        logger.error(f"数据库连接失败，无法记录日志: {operation_type}")
        return False

    try:
        cursor = connection.cursor()

        # 获取客户端IP地址
        try:
            import socket
            hostname = socket.gethostname()
            ip_address = socket.gethostbyname(hostname)
        except:
            ip_address = "127.0.0.1"

        # 如果有详细的记录数据，将其添加到操作详情中
        if record_data and isinstance(record_data, dict):
            details_with_data = f"{operation_details}\n\n记录详情：\n"

            # 添加记录类型
            if 'record_type' in record_data:
                details_with_data += f"• 记录类型：{record_data['record_type']}\n"

            # 添加所属人信息
            if 'owner' in record_data:
                details_with_data += f"• 所属人：{record_data['owner']}\n"

            # 添加基本信息
            if 'name' in record_data:
                details_with_data += f"• 姓名：{record_data['name']}\n"
            if 'amount' in record_data:
                details_with_data += f"• 金额：{record_data['amount']}元\n"
            if 'occasion' in record_data:
                details_with_data += f"• 事件：{record_data['occasion']}\n"
            if 'date' in record_data:
                details_with_data += f"• 日期：{record_data['date']}\n"

            # 添加回礼信息
            if 'return_amount' in record_data and record_data['return_amount'] and record_data['return_amount'] > 0:
                details_with_data += f"• 回礼金额：{record_data['return_amount']}元\n"
            if 'return_occasion' in record_data and record_data['return_occasion']:
                details_with_data += f"• 回礼事件：{record_data['return_occasion']}\n"
            if 'return_date' in record_data and record_data['return_date']:
                details_with_data += f"• 回礼日期：{record_data['return_date']}\n"

            # 添加备注信息
            if 'remark' in record_data and record_data['remark']:
                details_with_data += f"• 备注：{record_data['remark']}\n"

            operation_details = details_with_data

        # 记录新日志
        cursor.execute("""
            INSERT INTO system_logs (operation_type, operation_details, user_name, record_id, ip_address)
            VALUES (%s, %s, %s, %s, %s)
        """, (operation_type, operation_details, user_name, record_id, ip_address))

        # ===================== 新增：自动清理一周前的旧日志 =====================
        try:
            # 删除创建时间超过7天的日志
            cursor.execute("""
                DELETE FROM system_logs 
                WHERE created_at < DATE_SUB(NOW(), INTERVAL 7 DAY)
            """)
            deleted_count = cursor.rowcount
            
            if deleted_count > 0:
                logger.info(f"自动清理了 {deleted_count} 条一周前的旧日志")
                
                # 记录清理操作本身
                cursor.execute("""
                    INSERT INTO system_logs (operation_type, operation_details, user_name, ip_address)
                    VALUES (%s, %s, %s, %s)
                """, ("SYSTEM", f"自动清理日志 - 删除了{deleted_count}条一周前的旧日志", user_name, ip_address))
        except Error as e:
            logger.warning(f"清理旧日志时出错: {e}")
            # 不因清理失败而影响主要操作
        # ===================== 结束新增 =====================

        connection.commit()
        cursor.close()
        logger.info(f"操作日志记录成功: {operation_type}")
        return True
    except Error as e:
        logger.error(f"记录日志错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()

def calculate_completion_status(record):
    """计算完成状态"""
    record_type = record.get("record_type", "受礼记录")

    has_basic_info = (
            record.get("name") and
            record.get("amount", 0) > 0 and
            record.get("occasion") and
            record.get("date")
    )

    has_return_info = (
            record.get("return_amount", 0) > 0 and
            record.get("return_occasion") and
            record.get("return_date")
    )

    if record_type == "受礼记录":
        if has_basic_info and has_return_info:
            return "已完成"
        elif has_basic_info:
            return "仅受礼"
        else:
            return "未完成"
    elif record_type == "随礼记录":
        if has_basic_info and has_return_info:
            return "已完成"
        elif has_basic_info:
            return "仅随礼"
        else:
            return "未完成"
    return "未完成"

def get_pinyin_sort_key(name):
    """获取姓名的拼音排序键"""
    if not name:
        return ''
    
    # 如果安装了pypinyin，使用拼音排序
    if HAS_PINYIN:
        try:
            # 获取每个字的拼音首字母
            pinyin_list = pinyin(name, style=Style.FIRST_LETTER)
            # 转换为大写字符串用于排序
            return ''.join([p[0].upper() for p in pinyin_list if p])
        except Exception as e:
            logger.error(f"拼音转换错误: {e}")
    
    # 备用方案：使用Unicode编码排序（简单的中文排序）
    return name

def load_records():
    """从数据库加载记录"""
    connection = create_connection()
    if not connection:
        return []

    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, record_type, name, amount, occasion, date, 
                   has_returned, return_amount, return_occasion, return_date, remark, owner
            FROM gift_records 
            ORDER BY id DESC
        """)
        records = cursor.fetchall()
        cursor.close()

        for record in records:
            record['id'] = int(record['id'])
            record['amount'] = float(record['amount'])
            record['return_amount'] = float(record['return_amount'])
            record['has_returned'] = bool(record['has_returned'])
            
            # 格式化日期显示
            if record['date']:
                if isinstance(record['date'], str):
                    # 如果是字符串，直接使用
                    pass
                else:
                    # 如果是datetime对象，转换为字符串
                    record['date'] = record['date'].strftime("%Y-%m-%d")
            
            if record['return_date'] is None:
                record['return_date'] = ""
            elif record['return_date'] and not isinstance(record['return_date'], str):
                record['return_date'] = record['return_date'].strftime("%Y-%m-%d")
                
            if record['return_occasion'] is None:
                record['return_occasion'] = ""
            if record['remark'] is None:
                record['remark'] = ""
            if record['owner'] is None:
                record['owner'] = "郭宁"

        return records
    except Error as e:
        logger.error(f"加载记录错误: {e}")
        return []
    finally:
        if connection and connection.is_connected():
            connection.close()


def is_duplicate_gift_record(record, exclude_id=None):
    """检查礼尚往来记录是否重复"""
    connection = create_connection()
    if not connection:
        return False

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件 - 检查主要字段是否相同
        query = """
            SELECT id FROM gift_records 
            WHERE record_type = %s 
            AND name = %s 
            AND amount = %s 
            AND occasion = %s 
            AND date = %s 
            AND owner = %s
        """
        params = [
            record['record_type'],
            record['name'],
            record['amount'],
            record['occasion'],
            record['date'],
            record['owner']
        ]
        
        # 如果是更新操作，排除当前记录
        if exclude_id:
            query += " AND id != %s"
            params.append(exclude_id)
        
        cursor.execute(query, params)
        result = cursor.fetchone()
        
        return result is not None
        
    except Error as e:
        logger.error(f"检查重复礼尚往来记录错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()


def save_record(record):
    """保存单个记录到数据库"""
    connection = create_connection()
    if not connection:
        return False

    try:
        cursor = connection.cursor()

        return_date = record.get('return_date', '')
        if return_date == '':
            return_date = None

        date = record.get('date', '')
        if date == '':
            date = None

        record_type = record.get('record_type', '受礼记录')
        owner = record.get('owner', '郭宁')
        has_returned = bool(record.get('return_amount', 0) > 0 and
                            record.get('return_occasion') and
                            record.get('return_date'))

        is_update = 'id' in record and record['id']

        if is_update:
            cursor.execute("""
                UPDATE gift_records 
                SET record_type = %s, name = %s, amount = %s, occasion = %s, date = %s,
                    has_returned = %s, return_amount = %s, return_occasion = %s, 
                    return_date = %s, remark = %s, owner = %s
                WHERE id = %s
            """, (
                record_type, record['name'], record['amount'], record['occasion'], date,
                has_returned, record['return_amount'], record['return_occasion'],
                return_date, record['remark'], owner, record['id']
            ))
            operation_type = "EDIT"
            operation_details = f"修改{record_type}"
            record_id = record['id']
        else:
            cursor.execute("""
                INSERT INTO gift_records 
                (record_type, name, amount, occasion, date, has_returned, return_amount, return_occasion, return_date, remark, owner)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                record_type, record['name'], record['amount'], record['occasion'], date,
                has_returned, record['return_amount'], record['return_occasion'],
                return_date, record['remark'], owner
            ))
            operation_type = "ADD"
            operation_details = f"添加{record_type}"
            record_id = cursor.lastrowid

        connection.commit()
        cursor.close()

        # 记录操作日志
        log_operation(operation_type, operation_details, record_id, record_data=record)
        return True
    except Error as e:
        logger.error(f"保存记录错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()

def delete_record_by_id(record_id):
    """根据ID删除记录"""
    connection = create_connection()
    if not connection:
        return False

    try:
        # 先获取记录信息用于日志
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM gift_records WHERE id = %s", (record_id,))
        record = cursor.fetchone()

        cursor.execute("DELETE FROM gift_records WHERE id = %s", (record_id,))
        connection.commit()
        cursor.close()

        # 记录删除日志
        if record:
            log_operation("DELETE",
                          f"删除{record['record_type']}",
                          record_id,
                          record_data=record)

        return True
    except Error as e:
        logger.error(f"删除记录错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：记账相关函数 =====================
def is_duplicate_account_record(record, exclude_id=None):
    """检查记账记录是否重复"""
    connection = create_connection()
    if not connection:
        return False

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件 - 检查主要字段是否相同
        query = """
            SELECT id FROM daily_accounts 
            WHERE record_type = %s 
            AND category = %s 
            AND subcategory = %s 
            AND amount = %s 
            AND account_date = %s 
            AND owner = %s
        """
        params = [
            record['record_type'],
            record['category'],
            record['subcategory'] or '',  # 处理None值
            record['amount'],
            record['account_date'],
            record['owner']
        ]
        
        # 如果是更新操作，排除当前记录
        if exclude_id:
            query += " AND id != %s"
            params.append(exclude_id)
        
        cursor.execute(query, params)
        result = cursor.fetchone()
        
        return result is not None
        
    except Error as e:
        logger.error(f"检查重复记录错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()

def save_account_record(record):
    """保存记账记录到数据库"""
    connection = create_connection()
    if not connection:
        return False

    try:
        cursor = connection.cursor()

        account_date = record.get('account_date', '')
        if account_date == '':
            account_date = None

        is_update = 'id' in record and record['id']
        record_id = record.get('id')

        # 检查是否重复（更新时排除当前记录）
        if is_duplicate_account_record(record, exclude_id=record_id):
            logger.info(f"发现重复记录: {record}")
            return 'duplicate'

        if is_update:
            cursor.execute("""
                UPDATE daily_accounts 
                SET record_type = %s, category = %s, subcategory = %s, amount = %s, 
                    account_date = %s, description = %s, payment_method = %s, owner = %s
                WHERE id = %s
            """, (
                record['record_type'], record['category'], record['subcategory'], 
                record['amount'], account_date, record['description'], 
                record['payment_method'], record['owner'], record['id']
            ))
            operation_type = "EDIT"
            operation_details = f"修改记账记录 - 类别: {record['category']}, 金额: {record['amount']}"
            record_id = record['id']
        else:
            cursor.execute("""
                INSERT INTO daily_accounts 
                (record_type, category, subcategory, amount, account_date, description, payment_method, owner)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                record['record_type'], record['category'], record['subcategory'], 
                record['amount'], account_date, record['description'], 
                record['payment_method'], record['owner']
            ))
            operation_type = "ADD"
            operation_details = f"添加记账记录 - 类别: {record['category']}, 金额: {record['amount']}"
            record_id = cursor.lastrowid

        connection.commit()
        cursor.close()

        # 记录操作日志
        log_operation(operation_type, operation_details, record_id, record_data=record)
        return True
    except Error as e:
        logger.error(f"保存记账记录错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()

def delete_account_record_by_id(record_id):
    """根据ID删除记账记录"""
    connection = create_connection()
    if not connection:
        return False

    try:
        # 先获取记录信息用于日志
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM daily_accounts WHERE id = %s", (record_id,))
        record = cursor.fetchone()

        cursor.execute("DELETE FROM daily_accounts WHERE id = %s", (record_id,))
        connection.commit()
        cursor.close()

        # 记录删除日志
        if record:
            log_operation("DELETE",
                          f"删除记账记录 - 类别: {record['category']}, 金额: {record['amount']}",
                          record_id,
                          record_data=record)

        return True
    except Error as e:
        logger.error(f"删除记账记录错误: {e}")
        return False
    finally:
        if connection and connection.is_connected():
            connection.close()

def load_account_records():
    """从数据库加载记账记录"""
    connection = create_connection()
    if not connection:
        return []

    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, record_type, category, subcategory, amount, account_date, 
                   description, payment_method, owner
            FROM daily_accounts 
            ORDER BY account_date DESC, id DESC
        """)
        records = cursor.fetchall()
        cursor.close()

        for record in records:
            record['id'] = int(record['id'])
            record['amount'] = float(record['amount'])
            
            # 格式化日期显示
            if record['account_date']:
                if isinstance(record['account_date'], str):
                    pass
                else:
                    record['account_date'] = record['account_date'].strftime("%Y-%m-%d")
            
            if record['subcategory'] is None:
                record['subcategory'] = ""
            if record['description'] is None:
                record['description'] = ""
            if record['payment_method'] is None:
                record['payment_method'] = "现金"
            if record['owner'] is None:
                record['owner'] = "郭宁"

        return records
    except Error as e:
        logger.error(f"加载记账记录错误: {e}")
        return []
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 修复：高级统计功能 =====================
# 在app.py中找到get_account_statistics_by_period函数，修改以下部分：

# ===================== 修复：高级统计功能 =====================
def get_account_statistics_by_period(stat_type, start_date=None, end_date=None, owner=None):
    """按时间段统计记账数据（优化版）"""
    connection = create_connection()
    if not connection:
        return []

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if start_date:
            where_conditions.append("account_date >= %s")
            params.append(start_date)
        
        if end_date:
            where_conditions.append("account_date <= %s")
            params.append(end_date)
            
        if owner and owner != "全部":
            where_conditions.append("owner = %s")
            params.append(owner)
        
        where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""
        
        # 使用更高效的查询
        if stat_type == "monthly":
            # 修复按月统计查询 - 确保返回正确的字段名和格式
            query = f"""
                SELECT 
                    CONCAT(YEAR(account_date), '年', LPAD(MONTH(account_date), 2, '0'), '月') as period_name,
                    CONCAT(YEAR(account_date), LPAD(MONTH(account_date), 2, '0')) as sort_key,
                    record_type,
                    owner,
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                {where_clause}
                GROUP BY YEAR(account_date), MONTH(account_date), record_type, owner
                ORDER BY sort_key DESC, record_type, owner
            """
        elif stat_type == "quarterly":
            #按季度统计
            query = f"""
                SELECT 
                    CONCAT(YEAR(account_date), '年第', QUARTER(account_date), '季度') as period_name,
                    CONCAT(YEAR(account_date), LPAD(QUARTER(account_date), 2, '0')) as sort_key,
                    record_type,
                    owner,
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                {where_clause}
                GROUP BY YEAR(account_date), QUARTER(account_date), record_type, owner
                ORDER BY sort_key DESC, record_type, owner
            """
        elif stat_type == "yearly":
            #按年统计
            query = f"""
                SELECT 
                    CONCAT(YEAR(account_date), '年') as period_name,
                    YEAR(account_date) as sort_key,
                    record_type,
                    owner,
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                {where_clause}
                GROUP BY YEAR(account_date), record_type, owner
                ORDER BY sort_key DESC, record_type, owner
            """
        elif stat_type == "category":
            #按类别统计
            query = f"""
                SELECT 
                    category as period_name,
                    record_type,
                    owner,
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                {where_clause}
                GROUP BY category, record_type, owner
                ORDER BY category, record_type, owner
            """
        elif stat_type == "subcategory":
            #按子类别统计
            query = f"""
                SELECT 
                    CONCAT(category, '-', subcategory) as period_name,
                    record_type,
                    category,
                    subcategory,
                    owner,
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                {where_clause}
                GROUP BY category, subcategory, record_type, owner
                ORDER BY category, subcategory, record_type, owner
            """
        elif stat_type == "owner_detail":
            #按所属人详细统计
            query = f"""
                SELECT 
                    owner as period_name,
                    record_type,
                    category,
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                {where_clause}
                GROUP BY owner, record_type, category
                ORDER BY owner, record_type, category
            """
        else:
            return []
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # 格式化结果
        for result in results:
            result['total_amount'] = float(result['total_amount'])
        
        cursor.close()
        return results
        
    except Error as e:
        logger.error(f"统计查询错误: {e}")
        return []
    finally:
        if connection and connection.is_connected():
            connection.close()

def get_account_summary_statistics(start_date=None, end_date=None, owner=None):
    """获取记账数据汇总统计"""
    connection = create_connection()
    if not connection:
        return {
            'total': {'total_count': 0, 'total_expense': 0.0, 'total_income': 0.0},
            'by_owner': [],
            'by_category': []
        }

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if start_date:
            where_conditions.append("account_date >= %s")
            params.append(start_date)
        
        if end_date:
            where_conditions.append("account_date <= %s")
            params.append(end_date)
            
        if owner and owner != "全部":
            where_conditions.append("owner = %s")
            params.append(owner)
        
        where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""
        
        # 总统计
        total_query = f"""
            SELECT 
                COUNT(*) as total_count,
                COALESCE(SUM(CASE WHEN record_type = '支出' THEN amount ELSE 0 END), 0) as total_expense,
                COALESCE(SUM(CASE WHEN record_type = '收入' THEN amount ELSE 0 END), 0) as total_income
            FROM daily_accounts 
            {where_clause}
        """
        
        cursor.execute(total_query, params)
        total_stats = cursor.fetchone()
        
        # 确保所有字段都有默认值
        if total_stats:
            total_stats['total_count'] = total_stats.get('total_count', 0) or 0
            total_stats['total_expense'] = float(total_stats.get('total_expense', 0) or 0)
            total_stats['total_income'] = float(total_stats.get('total_income', 0) or 0)
        else:
            total_stats = {'total_count': 0, 'total_expense': 0.0, 'total_income': 0.0}
        
        # 按所属人统计
        owner_query = f"""
            SELECT 
                owner,
                COUNT(*) as count,
                COALESCE(SUM(CASE WHEN record_type = '支出' THEN amount ELSE 0 END), 0) as expense,
                COALESCE(SUM(CASE WHEN record_type = '收入' THEN amount ELSE 0 END), 0) as income
            FROM daily_accounts 
            {where_clause}
            GROUP BY owner
        """
        
        cursor.execute(owner_query, params)
        owner_stats = cursor.fetchall()
        
        # 按类别统计（前10个）
        category_query = f"""
            SELECT 
                record_type,
                category,
                COUNT(*) as count,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            {where_clause}
            GROUP BY record_type, category
            ORDER BY total_amount DESC
            LIMIT 10
        """
        
        cursor.execute(category_query, params)
        category_stats = cursor.fetchall()
        
        cursor.close()
        
        return {
            'total': total_stats,
            'by_owner': owner_stats,
            'by_category': category_stats
        }
        
    except Error as e:
        logger.error(f"汇总统计查询错误: {e}")
        return {
            'total': {'total_count': 0, 'total_expense': 0.0, 'total_income': 0.0},
            'by_owner': [],
            'by_category': []
        }
    finally:
        if connection and connection.is_connected():
            connection.close()

# Flask路由
@app.route('/')
def index():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('account_management.html')


# 添加礼尚往来记录管理路由
@app.route('/gift_management')
@login_required
def gift_management():
    """礼尚往来记录管理页面"""
    return render_template('index.html')

# 修改记账管理路由（保持兼容性）
@app.route('/account_management')
@login_required
def account_management():
    """记账管理页面（兼容旧链接）"""
    return render_template('account_management.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', 'admin')  # 默认为admin
        password = request.form.get('password')
        
        if verify_login(username, password):
            session['logged_in'] = True
            session['username'] = username
            session['login_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            session['last_activity'] = time.time()
            
            # 记录登录日志
            log_operation("LOGIN", f"用户登录系统 - 用户名: {username}", user_name=username)
            return jsonify({'success': True})
        else:
            logger.warning(f"登录失败: 用户名 '{username}' 密码错误")
            return jsonify({'success': False, 'message': '用户名或密码错误'})
    
    # GET请求时返回登录页面，并预填用户名（如果有的话）
    default_username = get_default_username()
    return render_template('login.html', default_username=default_username)

@app.route('/logout')
def logout():
    username = session.get('username', 'admin')
    # 记录退出日志
    log_operation("SYSTEM", "用户退出系统", user_name=username)
    session.clear()
    return redirect(url_for('login'))

# 忘记密码相关路由
@app.route('/api/forgot_password/verify_user', methods=['POST'])
def forgot_password_verify_user():
    """验证用户名是否存在并返回安全问题"""
    data = request.json
    username = data.get('username', '').strip()
    
    if not username:
        return jsonify({'success': False, 'message': '请输入用户名'})
    
    security_question = get_user_security_question(username)
    
    if not security_question:
        return jsonify({'success': False, 'message': '用户名不存在或未设置安全问题'})
    
    return jsonify({
        'success': True, 
        'security_question': security_question
    })

@app.route('/api/forgot_password/verify_answer', methods=['POST'])
def forgot_password_verify_answer():
    """验证安全问题答案"""
    data = request.json
    username = data.get('username', '').strip()
    answer = data.get('answer', '').strip()
    
    if not username or not answer:
        return jsonify({'success': False, 'message': '用户名和安全问题答案不能为空'})
    
    if verify_security_answer(username, answer):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': '安全问题答案错误'})

@app.route('/api/forgot_password/reset_password', methods=['POST'])
def forgot_password_reset_password():
    """重置密码"""
    data = request.json
    username = data.get('username', '').strip()
    new_password = data.get('new_password', '')
    confirm_password = data.get('confirm_password', '')
    
    if not username or not new_password or not confirm_password:
        return jsonify({'success': False, 'message': '所有字段都必须填写'})
    
    if new_password != confirm_password:
        return jsonify({'success': False, 'message': '两次输入的密码不一致'})
    
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': '密码长度不能少于6位'})
    
    if reset_user_password(username, new_password):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': '重置密码失败'})

# 修改安全问题路由
@app.route('/api/change_security_question', methods=['POST'])
@login_required
def change_security_question():
    """修改用户的安全问题和答案"""
    data = request.json
    username = session.get('username', 'admin')  # 从session获取当前登录用户
    new_question = data.get('new_question')
    new_answer = data.get('new_answer')
    password = data.get('password')  # 需要验证密码
    
    # 验证密码
    if not verify_login(username, password):
        return jsonify({'success': False, 'message': '密码错误'})
    
    if not new_question or not new_answer:
        return jsonify({'success': False, 'message': '安全问题和答案不能为空'})
    
    connection = create_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'})

    try:
        cursor = connection.cursor()
        
        # 加密新的安全答案
        answer_hash, answer_salt = encrypt_password(new_answer)
        
        cursor.execute("""
            UPDATE user_security 
            SET security_question = %s, 
                security_answer_hash = %s, 
                security_answer_salt = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE username = %s
        """, (new_question, answer_hash, answer_salt, username))
        
        connection.commit()
        cursor.close()
        
        # 记录操作日志
        log_operation("SYSTEM", f"修改安全问题 - 用户名: {username}, 新问题: {new_question}", user_name=username)
        
        return jsonify({'success': True, 'message': '安全问题修改成功'})
        
    except Error as e:
        logger.error(f"修改安全问题错误: {e}")
        return jsonify({'success': False, 'message': f'修改失败: {e}'})
    finally:
        if connection and connection.is_connected():
            connection.close()
            
@app.route('/api/event_statistics')
@login_required
def get_event_statistics():
    """获取事件金额统计（基于整个数据库）"""
    try:
        event_name = request.args.get('event_name', '').strip()
        
        if not event_name:
            return jsonify({'success': False, 'message': '事件名称不能为空'})
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 获取数据库中所有记录来计算事件统计
        cursor.execute("""
            SELECT id, record_type, name, amount, occasion, date, 
                   has_returned, return_amount, return_occasion, return_date, remark, owner
            FROM gift_records 
        """)
        all_records = cursor.fetchall()
        
        # 筛选相关记录
        related_records = []
        
        # 1. 筛选受礼记录中事件名称包含输入事件名称的记录
        gift_records = [r for r in all_records if 
                       r["record_type"] == "受礼记录" and 
                       r["occasion"] and event_name in r["occasion"]]
        
        # 2. 筛选随礼记录中回礼事件包含输入事件名称的记录
        return_records = [r for r in all_records if 
                         r["record_type"] == "随礼记录" and 
                         r["return_occasion"] and event_name in r["return_occasion"]]
        
        # 合并相关记录
        related_records.extend(gift_records)
        related_records.extend(return_records)
        
        if not related_records:
            cursor.close()
            return jsonify({
                'success': False, 
                'message': f'没有找到与"{event_name}"相关的记录'
            })
        
        # 计算受礼总额（受礼记录中的金额）
        gift_amount = sum(float(r["amount"]) for r in gift_records)
        
        # 计算回礼总额（随礼记录中的回礼金额）
        return_amount = sum(float(r.get("return_amount", 0)) for r in return_records)
        
        # 计算总金额
        total_amount = gift_amount + return_amount
        
        # 处理日期格式
        for record in related_records:
            if record['date'] and not isinstance(record['date'], str):
                record['date'] = record['date'].strftime("%Y-%m-%d")
            if record['return_date'] and not isinstance(record['return_date'], str):
                record['return_date'] = record['return_date'].strftime("%Y-%m-%d")
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'gift_amount': gift_amount,
            'return_amount': return_amount,
            'total_amount': total_amount,
            'records_count': len(related_records),
            'related_records': related_records
        })
        
    except Error as e:
        logger.error(f"获取事件统计错误: {e}")
        return jsonify({'success': False, 'message': '获取事件统计失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：回礼记录统计API =====================
@app.route('/api/return_records/statistics')
@login_required
def get_return_records_statistics():
    """获取回礼记录统计"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner', '全部')
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': '请选择开始日期和结束日期'})
        
        logger.info(f"回礼记录统计 - 开始日期: {start_date}, 结束日期: {end_date}, 所属人: {owner}")
        
        connection = create_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = ["date BETWEEN %s AND %s"]
        params = [start_date, end_date]
        
        # 添加所属人筛选
        if owner != "全部":
            where_conditions.append("owner = %s")
            params.append(owner)
        
        # 查询条件：回礼事件不为空
        # 包括：1. 随礼记录（回礼事件不为空） 2. 受礼记录（回礼事件不为空）
        where_conditions.append("(return_occasion IS NOT NULL AND return_occasion != '')")
        
        where_clause = " AND ".join(where_conditions)
        
        # 查询符合条件的记录
        query = f"""
            SELECT id, record_type, name, amount, occasion, date, 
                   has_returned, return_amount, return_occasion, return_date, remark, owner
            FROM gift_records 
            WHERE {where_clause}
            ORDER BY date DESC
        """
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # 计算总支出金额
        total_amount = 0.0
        for record in records:
            if record['record_type'] == '随礼记录':
                # 随礼记录：支出金额 = 金额
                total_amount += float(record['amount'])
            elif record['record_type'] == '受礼记录' and record['return_amount']:
                # 受礼记录：支出金额 = 回礼金额
                total_amount += float(record['return_amount'] or 0)
        
        # 格式化记录数据
        formatted_records = []
        for record in records:
            formatted_record = {
                'id': record['id'],
                'record_type': record['record_type'],
                'name': record['name'],
                'amount': float(record['amount']),
                'occasion': record['occasion'],
                'date': record['date'].strftime('%Y-%m-%d') if record['date'] else '',
                'has_returned': bool(record['has_returned']),
                'return_amount': float(record['return_amount'] or 0),
                'return_occasion': record['return_occasion'] or '',
                'return_date': record['return_date'].strftime('%Y-%m-%d') if record['return_date'] else '',
                'owner': record['owner'] or '郭宁'
            }
            formatted_records.append(formatted_record)
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'records_count': len(records),
            'total_amount': total_amount,
            'records': formatted_records,
            'query_params': {
                'start_date': start_date,
                'end_date': end_date,
                'owner': owner
            }
        })
        
    except Error as e:
        logger.error(f"获取回礼记录统计错误: {e}")
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"回礼记录统计异常: {str(e)}")
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'}), 500

@app.route('/api/return_records/statistics/export')
@login_required
def export_return_records_statistics():
    """导出回礼记录统计结果"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner', '全部')
        
        if not start_date or not end_date:
            return jsonify({'error': '请选择开始日期和结束日期'}), 400
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = ["date BETWEEN %s AND %s"]
        params = [start_date, end_date]
        
        if owner != "全部":
            where_conditions.append("owner = %s")
            params.append(owner)
        
        where_conditions.append("(return_occasion IS NOT NULL AND return_occasion != '')")
        where_clause = " AND ".join(where_conditions)
        
        # 查询记录
        query = f"""
            SELECT record_type, name, amount, occasion, date, 
                   return_amount, return_occasion, return_date, owner
            FROM gift_records 
            WHERE {where_clause}
            ORDER BY date DESC
        """
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # 计算总支出金额
        total_amount = 0.0
        expense_by_type = {'随礼记录': 0.0, '受礼记录': 0.0}
        for record in records:
            if record['record_type'] == '随礼记录':
                amount = float(record['amount'])
                total_amount += amount
                expense_by_type['随礼记录'] += amount
            elif record['record_type'] == '受礼记录' and record['return_amount']:
                amount = float(record['return_amount'] or 0)
                total_amount += amount
                expense_by_type['受礼记录'] += amount
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "回礼记录统计"
        
        # 设置样式
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # 写入标题
        title = f"回礼记录统计报告 - {start_date} 至 {end_date}"
        if owner != "全部":
            title += f" (所属人: {owner})"
        
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:I1')
        
        # 写入汇总信息
        ws['A3'] = "汇总统计"
        ws['A3'].font = header_font
        
        ws['A4'] = "总记录数"
        ws['B4'] = len(records)
        
        ws['A5'] = "总支出金额"
        ws['B5'] = total_amount
        
        ws['A6'] = "随礼记录支出"
        ws['B6'] = expense_by_type['随礼记录']
        
        ws['A7'] = "受礼记录回礼支出"
        ws['B7'] = expense_by_type['受礼记录']
        
        # 写入表头
        headers = ['记录类型', '姓名', '事件', '金额', '日期', '回礼金额', '回礼事件', '回礼日期', '所属人']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row, record in enumerate(records, 10):
            # 根据记录类型计算支出金额
            expense_amount = record['amount']
            if record['record_type'] == '受礼记录' and record['return_amount']:
                expense_amount = record['return_amount']
            
            ws.cell(row=row, column=1, value=record['record_type']).border = border
            ws.cell(row=row, column=2, value=record['name']).border = border
            ws.cell(row=row, column=3, value=record['occasion']).border = border
            ws.cell(row=row, column=4, value=float(expense_amount)).border = border
            ws.cell(row=row, column=5, value=record['date'].strftime('%Y-%m-%d') if not isinstance(record['date'], str) else record['date']).border = border
            ws.cell(row=row, column=6, value=float(record['return_amount'] or 0) if record['return_amount'] else '-').border = border
            ws.cell(row=row, column=7, value=record['return_occasion'] or '-').border = border
            ws.cell(row=row, column=8, value=record['return_date'].strftime('%Y-%m-%d') if record['return_date'] and not isinstance(record['return_date'], str) else (record['return_date'] or '-')).border = border
            ws.cell(row=row, column=9, value=record['owner'] or '郭宁').border = border
        
        # 设置列宽
        column_widths = [12, 12, 20, 12, 12, 12, 15, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 设置金额格式
        for row in range(10, len(records) + 10):
            ws.cell(row=row, column=4).number_format = '0.00'
            ws.cell(row=row, column=6).number_format = '0.00'
        
        # 汇总金额格式
        for row in range(5, 8):
            ws.cell(row=row, column=2).number_format = '0.00'
        
        cursor.close()
        connection.close()
        
        # 创建内存文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"回礼记录统计_{start_date}_至_{end_date}"
        if owner != "全部":
            filename += f"_{owner}"
        filename += ".xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"导出回礼记录统计错误: {str(e)}")
        return jsonify({'error': '导出失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# 修改加载记录函数，添加分页
@app.route('/api/records')
@login_required
def get_records():
    """获取记录（带分页）"""
    try:
        # 获取分页参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        sort_method = request.args.get('sort_method', '按记录类型排序')
        
        logger.info(f"获取记录请求 - 页码: {page}, 每页: {per_page}, 排序方式: {sort_method}")
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 获取所有记录
        cursor.execute("""
            SELECT id, record_type, name, amount, occasion, date, 
                   has_returned, return_amount, return_occasion, return_date, remark, owner
            FROM gift_records 
        """)
        all_records = cursor.fetchall()
        
        # 获取总记录数
        total = len(all_records)
        
        cursor.close()

        # 根据排序方式排序
        if sort_method == '按记录类型排序':
            all_records.sort(key=lambda x: (x['record_type'], x['date']), reverse=True)
        elif sort_method == '按姓名首字母排序':
            if HAS_PINYIN:
                from pypinyin import lazy_pinyin
                
                def get_pinyin_sort_key(name):
                    if not name:
                        return ''
                    try:
                        pinyin_list = lazy_pinyin(name, style=Style.FIRST_LETTER)
                        return ''.join([p[0].upper() for p in pinyin_list if p])
                    except Exception as e:
                        logger.error(f"拼音转换错误: {e}")
                        return name
                
                all_records.sort(key=lambda x: get_pinyin_sort_key(x['name']))
            else:
                all_records.sort(key=lambda x: x['name'] or '')
        elif sort_method == '按时间降序':
            all_records.sort(key=lambda x: x['date'] or '', reverse=True)
        elif sort_method == '按金额降序':
            all_records.sort(key=lambda x: float(x['amount']), reverse=True)
        
        # 分页处理
        start_idx = (page - 1) * per_page
        end_idx = min(start_idx + per_page, total)
        paginated_records = all_records[start_idx:end_idx]
        
        # 处理记录数据
        for record in paginated_records:
            record['id'] = int(record['id'])
            record['amount'] = float(record['amount'])
            record['return_amount'] = float(record['return_amount'] or 0)
            record['has_returned'] = bool(record['has_returned'])
            
            # 格式化日期显示
            if record['date']:
                if isinstance(record['date'], str):
                    pass
                else:
                    record['date'] = record['date'].strftime("%Y-%m-%d")
            
            if record['return_date'] is None:
                record['return_date'] = ""
            elif record['return_date'] and not isinstance(record['return_date'], str):
                record['return_date'] = record['return_date'].strftime("%Y-%m-%d")
                
            if record['return_occasion'] is None:
                record['return_occasion'] = ""
            if record['remark'] is None:
                record['remark'] = ""
            if record['owner'] is None:
                record['owner'] = "郭宁"

        # 计算总页数
        total_pages = (total + per_page - 1) // per_page if total > 0 else 1
        
        logger.info(f"获取记录响应 - 总记录数: {total}, 返回记录数: {len(paginated_records)}, 总页数: {total_pages}")

        return jsonify({
            'records': paginated_records,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages
        })
    except Error as e:
        logger.error(f"加载记录错误: {e}")
        return jsonify({'error': '加载记录失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/records/search', methods=['POST'])
@login_required
def search_records():
    """搜索记录（带分页）"""
    try:
        data = request.json
        
        # 获取分页参数
        page = data.get('page', 1)
        per_page = data.get('per_page', 50)
        
        logger.info(f"搜索记录请求 - 页码: {page}, 每页: {per_page}, 搜索条件: {data}")
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 构建查询
        query = """
            SELECT id, record_type, name, amount, occasion, date, 
                   has_returned, return_amount, return_occasion, return_date, remark, owner
            FROM gift_records 
            WHERE 1=1
        """
        count_query = "SELECT COUNT(*) as total FROM gift_records WHERE 1=1"
        params = []

        # 记录类型筛选
        record_type_filter = data.get('record_type', '全部')
        if record_type_filter != '全部':
            query += " AND record_type = %s"
            count_query += " AND record_type = %s"
            params.append(record_type_filter)

        # 姓名筛选
        name_filter = data.get('name', '').strip()
        if name_filter:
            query += " AND name LIKE %s"
            count_query += " AND name LIKE %s"
            params.append(f"%{name_filter}%")

        # 日期筛选
        date_filter = data.get('date', '').strip()
        if date_filter:
            query += " AND date = %s"
            count_query += " AND date = %s"
            params.append(date_filter)

        # 所属人筛选
        owner_filter = data.get('owner', '全部')
        if owner_filter != '全部':
            query += " AND owner = %s"
            count_query += " AND owner = %s"
            params.append(owner_filter)
        
        # 获取排序方式
        sort_method = data.get('sort_method', '按记录类型排序')
        
        # 根据排序方式添加ORDER BY子句
        if sort_method == '按记录类型排序':
            query += " ORDER BY record_type, date DESC"
        elif sort_method == '按姓名首字母排序':
            # 使用拼音首字母进行排序
            if HAS_PINYIN:
                # 如果安装了pypinyin，我们可以在应用层面排序
                # 先获取所有数据，然后在应用层面排序
                pass
            else:
                # 如果没有安装pypinyin，使用简单的Unicode排序
                query += " ORDER BY name"
        elif sort_method == '按时间降序':
            query += " ORDER BY date DESC"
        elif sort_method == '按金额降序':
            query += " ORDER BY amount DESC"
        
        # 获取总数
        cursor.execute(count_query, params)
        total_result = cursor.fetchone()
        total = total_result['total'] if total_result else 0
        
        # 执行查询获取记录
        cursor.execute(query, params)
        all_records = cursor.fetchall()
        
        # 完成状态筛选（在应用层面处理）
        status_filter = data.get('completion_status', '全部')
        if status_filter != '全部':
            filtered_records = []
            for record in all_records:
                current_status = calculate_completion_status(record)
                if status_filter == current_status:
                    filtered_records.append(record)
            all_records = filtered_records
            total = len(all_records)
        
        # 如果需要按拼音排序且安装了pypinyin
        if sort_method == '按姓名首字母排序' and HAS_PINYIN:
            from pypinyin import lazy_pinyin
            
            def get_pinyin_sort_key(name):
                if not name:
                    return ''
                # 获取每个字的拼音，取首字母
                try:
                    pinyin_list = lazy_pinyin(name, style=Style.FIRST_LETTER)
                    # 转换为大写字符串用于排序
                    return ''.join([p[0].upper() for p in pinyin_list if p])
                except Exception as e:
                    logger.error(f"拼音转换错误: {e}")
                    return name
            
            # 按拼音首字母排序
            all_records.sort(key=lambda x: get_pinyin_sort_key(x['name']))
        
        # 分页处理
        start_idx = (page - 1) * per_page
        end_idx = min(start_idx + per_page, total)
        paginated_records = all_records[start_idx:end_idx]
        
        cursor.close()

        # 处理记录数据
        for record in paginated_records:
            record['id'] = int(record['id'])
            record['amount'] = float(record['amount'])
            record['return_amount'] = float(record['return_amount'] or 0)
            record['has_returned'] = bool(record['has_returned'])
            
            # 格式化日期显示
            if record['date']:
                if isinstance(record['date'], str):
                    pass
                else:
                    record['date'] = record['date'].strftime("%Y-%m-%d")
            
            if record['return_date'] is None:
                record['return_date'] = ""
            elif record['return_date'] and not isinstance(record['return_date'], str):
                record['return_date'] = record['return_date'].strftime("%Y-%m-%d")
                
            if record['return_occasion'] is None:
                record['return_occasion'] = ""
            if record['remark'] is None:
                record['remark'] = ""
            if record['owner'] is None:
                record['owner'] = "郭宁"

        # 计算总页数
        total_pages = (total + per_page - 1) // per_page if total > 0 else 1
        
        logger.info(f"搜索记录响应 - 总记录数: {total}, 返回记录数: {len(paginated_records)}, 总页数: {total_pages}")

        return jsonify({
            'records': paginated_records,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages
        })
        
    except Error as e:
        logger.error(f"搜索记录错误: {e}")
        return jsonify({'error': '搜索失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()
    
    

@app.route('/api/records', methods=['POST'])
@login_required
def add_record():
    try:
        data = request.json
        
        # 数据验证
        if not data.get('name') or not data.get('name').strip():
            return jsonify({'success': False, 'message': '姓名不能为空'})
        
        try:
            amount = float(data['amount'])
            if amount <= 0:
                return jsonify({'success': False, 'message': '金额必须大于0'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '金额格式错误'})
            
        if not data.get('occasion') or not data.get('occasion').strip():
            return jsonify({'success': False, 'message': '事件不能为空'})
            
        if not data.get('date') or not data.get('date').strip():
            return jsonify({'success': False, 'message': '日期不能为空'})
        
        # 处理回礼金额
        return_amount = 0.0
        if data.get('return_amount'):
            try:
                return_amount = float(data['return_amount'])
                if return_amount < 0:
                    return jsonify({'success': False, 'message': '回礼金额不能为负数'})
            except (ValueError, TypeError):
                return_amount = 0.0
        
        record = {
            'record_type': data.get('record_type', '受礼记录'),
            'owner': data.get('owner', '郭宁'),
            'name': data['name'].strip(),
            'amount': amount,
            'occasion': data['occasion'].strip(),
            'date': data['date'].strip(),
            'return_amount': return_amount,
            'return_occasion': data.get('return_occasion', '').strip(),
            'return_date': data.get('return_date', '').strip(),
            'remark': data.get('remark', '').strip()
        }
        
        # ==================== 新增：检查重复记录 ====================
        if is_duplicate_gift_record(record):
            return jsonify({
                'success': False, 
                'message': '该记录已存在，请勿重复添加！',
                'duplicate': True  # 新增标识，用于前端区分错误类型
            })
        # ==================== 结束新增 ====================
        
        if save_record(record):
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': '保存到数据库失败'})
            
    except Exception as e:
        logger.error(f"添加记录错误: {str(e)}")
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})
        
        

@app.route('/api/records/<int:record_id>', methods=['PUT'])
@login_required
def update_record(record_id):
    try:
        data = request.json
        
        # 数据验证
        if not data.get('name') or not data.get('name').strip():
            return jsonify({'success': False, 'message': '姓名不能为空'})
        
        try:
            amount = float(data['amount'])
            if amount <= 0:
                return jsonify({'success': False, 'message': '金额必须大于0'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '金额格式错误'})
            
        if not data.get('occasion') or not data.get('occasion').strip():
            return jsonify({'success': False, 'message': '事件不能为空'})
            
        if not data.get('date') or not data.get('date').strip():
            return jsonify({'success': False, 'message': '日期不能为空'})
        
        # 处理回礼金额
        return_amount = 0.0
        if data.get('return_amount'):
            try:
                return_amount = float(data['return_amount'])
                if return_amount < 0:
                    return jsonify({'success': False, 'message': '回礼金额不能为负数'})
            except (ValueError, TypeError):
                return_amount = 0.0
        
        record = {
            'id': record_id,
            'record_type': data.get('record_type', '受礼记录'),
            'owner': data.get('owner', '郭宁'),
            'name': data['name'].strip(),
            'amount': amount,
            'occasion': data['occasion'].strip(),
            'date': data['date'].strip(),
            'return_amount': return_amount,
            'return_occasion': data.get('return_occasion', '').strip(),
            'return_date': data.get('return_date', '').strip(),
            'remark': data.get('remark', '').strip()
        }
        
        # ==================== 新增：检查重复记录（排除当前记录） ====================
        if is_duplicate_gift_record(record, exclude_id=record_id):
            return jsonify({
                'success': False, 
                'message': '该记录已存在，请勿重复添加！',
                'duplicate': True
            })
        # ==================== 结束新增 ====================
        
        if save_record(record):
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': '更新数据库失败'})
            
    except Exception as e:
        logger.error(f"更新记录错误: {str(e)}")
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})

@app.route('/api/records/<int:record_id>', methods=['DELETE'])
@login_required
def delete_record(record_id):
    if delete_record_by_id(record_id):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': '删除失败'})

@app.route('/api/statistics')
@login_required
def get_statistics():
    """获取统计数据（基于整个数据库，而不是当前页）"""
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 从数据库中获取所有记录来计算统计
        cursor.execute("""
            SELECT id, record_type, name, amount, occasion, date, 
                   has_returned, return_amount, return_occasion, return_date, remark, owner
            FROM gift_records 
        """)
        all_records = cursor.fetchall()
        
        # 基础统计
        gift_records = [r for r in all_records if r["record_type"] == "受礼记录"]
        return_records = [r for r in all_records if r["record_type"] == "随礼记录"]
        
        # 按所属人统计
        gift_records_a = [r for r in gift_records if r.get('owner') == '郭宁' or not r.get('owner')]
        return_records_a = [r for r in return_records if r.get('owner') == '郭宁' or not r.get('owner')]
        gift_records_b = [r for r in gift_records if r.get('owner') == '李佳慧']
        return_records_b = [r for r in return_records if r.get('owner') == '李佳慧']
        
        # ==================== 按照新规则计算金额 ====================
        # 郭宁受礼总额 = 郭宁的受礼记录金额 + 郭宁的随礼记录中的回礼金额
        total_gift_amount_a = float(0)
        total_return_amount_a = float(0)
        total_gift_amount_b = float(0)
        total_return_amount_b = float(0)
        
        # 确保金额转换为浮点数
        for r in gift_records_a:
            total_gift_amount_a += float(r["amount"] or 0)
        
        for r in return_records_a:
            total_gift_amount_a += float(r["return_amount"] or 0)
        
        for r in gift_records_a:
            total_return_amount_a += float(r["return_amount"] or 0)
        
        for r in return_records_a:
            total_return_amount_a += float(r["amount"] or 0)
        
        for r in gift_records_b:
            total_gift_amount_b += float(r["amount"] or 0)
        
        for r in return_records_b:
            total_gift_amount_b += float(r["return_amount"] or 0)
        
        for r in gift_records_b:
            total_return_amount_b += float(r["return_amount"] or 0)
        
        for r in return_records_b:
            total_return_amount_b += float(r["amount"] or 0)
        
        # 已完成回礼统计
        completed_records = [r for r in all_records if calculate_completion_status(r) == "已完成"]
        
        stats = {
            'total_count': len(all_records),
            'gift_count_a': len(gift_records_a),
            'return_count_a': len(return_records_a),
            'gift_count_b': len(gift_records_b),
            'return_count_b': len(return_records_b),
            'total_gift_amount_a': total_gift_amount_a,
            'total_return_amount_a': total_return_amount_a,
            'total_gift_amount_b': total_gift_amount_b,
            'total_return_amount_b': total_return_amount_b,
            'completed_count': len(completed_records)
        }
        
        cursor.close()
        logger.info(f"统计数据 - 总记录数: {len(all_records)}, 郭宁受礼金额: {total_gift_amount_a}, 李佳慧受礼金额: {total_gift_amount_b}")
        return jsonify(stats)
    except Error as e:
        logger.error(f"获取统计错误: {e}")
        return jsonify({'error': '获取统计失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/logs')
@login_required
def get_system_logs():
    # 获取查询参数
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    operation_type = request.args.get('operation_type', '')
    date_range = request.args.get('date_range', '')
    keyword = request.args.get('keyword', '')
    
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        query = """
            SELECT id, operation_type, operation_details, user_name, record_id, ip_address, created_at
            FROM system_logs 
            WHERE 1=1
        """
        params = []

        if operation_type and operation_type != '全部':
            query += " AND operation_type = %s"
            params.append(operation_type)

        if date_range and date_range != '全部':
            if date_range == '今天':
                query += " AND DATE(created_at) = CURDATE()"
            elif date_range == '最近7天':
                query += " AND created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)"
            elif date_range == '最近30天':
                query += " AND created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)"
            elif date_range == '最近3个月':
                query += " AND created_at >= DATE_SUB(NOW(), INTERVAL 3 MONTH)"

        if keyword:
            query += " AND (operation_details LIKE %s OR user_name LIKE %s)"
            params.extend([f"%{keyword}%", f"%{keyword}%"])

        # 获取总数
        count_query = f"SELECT COUNT(*) as total FROM ({query}) as t"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']

        # 添加排序和分页
        query += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
        offset = (page - 1) * per_page
        params.extend([per_page, offset])

        cursor.execute(query, params)
        logs = cursor.fetchall()
        
        # 格式化日期
        for log in logs:
            if log['created_at']:
                log['created_at'] = log['created_at'].strftime("%Y-%m-%d %H:%M:%S")

        cursor.close()
        
        return jsonify({
            'logs': logs,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })
        
    except Error as e:
        logger.error(f"加载日志错误: {e}")
        return jsonify({'error': '加载日志失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/change_password', methods=['POST'])
@login_required
def change_password():
    data = request.json
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    
    username = session.get('username', 'admin')
    
    if not verify_login(username, old_password):
        return jsonify({'success': False, 'message': '原密码错误'})
    
    if not new_password or len(new_password) < 6:
        return jsonify({'success': False, 'message': '新密码长度至少6位'})
    
    if reset_user_password(username, new_password):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': '修改密码失败'})

@app.route('/api/user_info')
@login_required
def get_user_info():
    """获取当前登录用户信息"""
    username = session.get('username', 'admin')
    login_time = session.get('login_time')
    return jsonify({
        'username': username,
        'login_time': login_time
    })

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    """获取所有用户列表"""
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor()
        cursor.execute("SELECT username FROM user_security")
        users = cursor.fetchall()
        cursor.close()
        
        user_list = [user[0] for user in users]
        return jsonify({'users': user_list})
        
    except Error as e:
        logger.error(f"获取用户列表错误: {e}")
        return jsonify({'error': '获取用户列表失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：记账管理API路由 =====================
@app.route('/api/account/categories')
@login_required
def get_account_categories():
    """获取记账类别"""
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        try:
            cursor.execute("""
                SELECT category_type, category_name, subcategories, sort_order 
                FROM account_categories 
                ORDER BY category_type, sort_order
            """)
            categories = cursor.fetchall()
        except Exception as e:
            logger.warning(f"查询account_categories表失败: {e}")
            # 如果没有类别数据，返回默认类别
            categories = [
                {'category_type': '支出', 'category_name': '食品酒水', 'subcategories': '["早餐", "午餐", "晚餐", "零食", "水果"]', 'sort_order': 0},
                {'category_type': '支出', 'category_name': '衣服饰品', 'subcategories': '["衣服", "裤子", "鞋子", "饰品"]', 'sort_order': 1},
                {'category_type': '支出', 'category_name': '居家物业', 'subcategories': '["房租", "水电费", "物业费"]', 'sort_order': 2},
                {'category_type': '收入', 'category_name': '工资收入', 'subcategories': '["工资", "奖金", "津贴"]', 'sort_order': 0},
                {'category_type': '收入', 'category_name': '投资收益', 'subcategories': '["股票", "基金", "理财"]', 'sort_order': 1}
            ]
        
        # 解析JSON字段
        for category in categories:
            if category.get('subcategories'):
                try:
                    category['subcategories'] = json.loads(category['subcategories'])
                except (json.JSONDecodeError, TypeError):
                    category['subcategories'] = []
            else:
                category['subcategories'] = []
        
        return jsonify(categories)
    except Error as e:
        logger.error(f"获取记账类别错误: {e}")
        # 返回默认类别
        default_categories = [
            {'category_type': '支出', 'category_name': '食品酒水', 'subcategories': ['早餐', '午餐', '晚餐', '零食', '水果'], 'sort_order': 0},
            {'category_type': '支出', 'category_name': '衣服饰品', 'subcategories': ['衣服', '裤子', '鞋子', '饰品'], 'sort_order': 1},
            {'category_type': '支出', 'category_name': '居家物业', 'subcategories': ['房租', '水电费', '物业费'], 'sort_order': 2},
            {'category_type': '收入', 'category_name': '工资收入', 'subcategories': ['工资', '奖金', '津贴'], 'sort_order': 0},
            {'category_type': '收入', 'category_name': '投资收益', 'subcategories': ['股票', '基金', '理财'], 'sort_order': 1}
        ]
        return jsonify(default_categories)
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/account/records')
@login_required
def get_account_records():
    """获取记账记录（带分页）"""
    try:
        # 获取分页参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 计算偏移量
        offset = (page - 1) * per_page
        
        # 查询当前页的记录
        cursor.execute("""
            SELECT id, record_type, category, subcategory, amount, account_date, 
                   description, payment_method, owner
            FROM daily_accounts 
            ORDER BY account_date DESC, id DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))
        records = cursor.fetchall()
        
        # 获取总记录数（使用估算值提高性能）
        cursor.execute("""
            SELECT COUNT(*) as total_count FROM daily_accounts
        """)
        total_result = cursor.fetchone()
        total_count = total_result['total_count'] if total_result else 0
        
        # 获取统计信息（优化版，避免多次查询）
        cursor.execute("""
            SELECT 
                COUNT(*) as total_count,
                COALESCE(SUM(CASE WHEN record_type = '支出' THEN amount ELSE 0 END), 0) as total_expense,
                COALESCE(SUM(CASE WHEN record_type = '收入' THEN amount ELSE 0 END), 0) as total_income
            FROM daily_accounts
        """)
        stats = cursor.fetchone()
        
        for record in records:
            record['id'] = int(record['id'])
            record['amount'] = float(record['amount'])
            if record['account_date'] and not isinstance(record['account_date'], str):
                record['account_date'] = record['account_date'].strftime("%Y-%m-%d")
            if record['subcategory'] is None:
                record['subcategory'] = ""
            if record['description'] is None:
                record['description'] = ""
            if record['payment_method'] is None:
                record['payment_method'] = "现金"
            if record['owner'] is None:
                record['owner'] = "郭宁"

        cursor.close()
        
        return jsonify({
            'records': records,
            'pagination': {
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': (total_count + per_page - 1) // per_page
            },
            'stats': {
                'total_count': stats['total_count'],
                'total_expense': float(stats['total_expense']),
                'total_income': float(stats['total_income']),
                'net_amount': float(stats['total_income']) - float(stats['total_expense'])
            }
        })
        
    except Error as e:
        logger.error(f"获取记账记录错误: {e}")
        return jsonify({'error': '获取记录失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/account/records', methods=['POST'])
@login_required
def add_account_record():
    """添加记账记录"""
    try:
        data = request.json
        
        # 数据验证
        if not data.get('category') or not data.get('category').strip():
            return jsonify({'success': False, 'message': '类别不能为空'})
        
        try:
            amount = float(data['amount'])
            if amount <= 0:
                return jsonify({'success': False, 'message': '金额必须大于0'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '金额格式错误'})
            
        if not data.get('account_date') or not data.get('account_date').strip():
            return jsonify({'success': False, 'message': '日期不能为空'})
        
        record = {
            'record_type': data.get('record_type', '支出'),
            'owner': data.get('owner', '郭宁'),
            'category': data['category'].strip(),
            'subcategory': data.get('subcategory', '').strip(),
            'amount': amount,
            'account_date': data['account_date'].strip(),
            'description': data.get('description', '').strip(),
            'payment_method': data.get('payment_method', '现金')
        }
        
        # 检查是否重复
        if is_duplicate_account_record(record):
            return jsonify({
                'success': False, 
                'message': '该记录已存在，请勿重复添加！',
                'duplicate': True
            })
        
        result = save_account_record(record)
        if result == 'duplicate':
            return jsonify({
                'success': False, 
                'message': '该记录已存在，请勿重复添加！',
                'duplicate': True
            })
        elif result:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': '保存到数据库失败'})
            
    except Exception as e:
        logger.error(f"添加记账记录错误: {str(e)}")
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})

@app.route('/api/account/records/<int:record_id>', methods=['PUT'])
@login_required
def update_account_record(record_id):
    """更新记账记录"""
    try:
        data = request.json
        
        # 数据验证
        if not data.get('category') or not data.get('category').strip():
            return jsonify({'success': False, 'message': '类别不能为空'})
        
        try:
            amount = float(data['amount'])
            if amount <= 0:
                return jsonify({'success': False, 'message': '金额必须大于0'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '金额格式错误'})
            
        if not data.get('account_date') or not data.get('account_date').strip():
            return jsonify({'success': False, 'message': '日期不能为空'})
        
        record = {
            'id': record_id,
            'record_type': data.get('record_type', '支出'),
            'owner': data.get('owner', '郭宁'),
            'category': data['category'].strip(),
            'subcategory': data.get('subcategory', '').strip(),
            'amount': amount,
            'account_date': data['account_date'].strip(),
            'description': data.get('description', '').strip(),
            'payment_method': data.get('payment_method', '现金')
        }
        
        # 检查是否重复（排除当前记录）
        if is_duplicate_account_record(record, exclude_id=record_id):
            return jsonify({
                'success': False, 
                'message': '该记录已存在，请勿重复添加！',
                'duplicate': True
            })
        
        result = save_account_record(record)
        if result == 'duplicate':
            return jsonify({
                'success': False, 
                'message': '该记录已存在，请勿重复添加！',
                'duplicate': True
            })
        elif result:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': '更新数据库失败'})
            
    except Exception as e:
        logger.error(f"更新记账记录错误: {str(e)}")
        return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})

@app.route('/api/account/records/<int:record_id>', methods=['DELETE'])
@login_required
def delete_account_record(record_id):
    """删除记账记录"""
    if delete_account_record_by_id(record_id):
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': '删除失败'})

# ===================== 修复：基础记账统计API =====================
@app.route('/api/account/statistics')
@login_required
def get_account_statistics():
    """获取记账统计信息"""
    try:
        records = load_account_records()
        
        # 基础统计
        expense_records = [r for r in records if r["record_type"] == "支出"]
        income_records = [r for r in records if r["record_type"] == "收入"]
        
        total_expense = sum(float(r["amount"]) for r in expense_records)
        total_income = sum(float(r["amount"]) for r in income_records)
        
        stats = {
            'basic': {
                'total_count': len(records),
                'expense_count': len(expense_records),
                'income_count': len(income_records),
                'total_expense': total_expense,
                'total_income': total_income,
                'net_amount': total_income - total_expense
            }
        }
        
        return jsonify(stats)
    except Exception as e:
        logger.error(f"获取记账统计错误: {str(e)}")
        return jsonify({
            'basic': {
                'total_count': 0,
                'expense_count': 0,
                'income_count': 0,
                'total_expense': 0.0,
                'total_income': 0.0,
                'net_amount': 0.0
            }
        })

# ===================== 新增：日历视图API =====================
@app.route('/api/account/calendar')
@login_required
def get_calendar_data():
    """获取日历视图数据"""
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        owner = request.args.get('owner', '全部')
        
        if not year or not month:
            # 如果没有提供年月，使用当前年月
            today = datetime.now()
            year = today.year
            month = today.month
        
        logger.info(f"获取日历数据 - 年份: {year}, 月份: {month}, 所属人: {owner}")
        
        # 计算该月的第一天和最后一天
        if month == 12:
            next_year = year + 1
            next_month = 1
        else:
            next_year = year
            next_month = month + 1
        
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{next_year}-{next_month:02d}-01"
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = ["account_date >= %s AND account_date < %s"]
        params = [start_date, end_date]
        
        if owner and owner != "全部":
            where_conditions.append("owner = %s")
            params.append(owner)
        
        where_clause = " AND ".join(where_conditions)
        
        # 查询每天每个所属人的收支数据
        query = f"""
            SELECT 
                account_date,
                owner,
                record_type,
                COUNT(*) as count,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            WHERE {where_clause}
            GROUP BY account_date, owner, record_type
            ORDER BY account_date, owner
        """
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # 获取该月所有日期
        calendar_data = {}
        current_date = datetime(year, month, 1)
        
        while current_date.month == month:
            date_str = current_date.strftime('%Y-%m-%d')
            calendar_data[date_str] = {
                'date': date_str,
                'day': current_date.day,
                'weekday': current_date.strftime('%a'),  # 星期几的缩写
                'owners': {}
            }
            current_date += timedelta(days=1)
        
        # 填充数据
        for record in results:
            date_str = record['account_date'].strftime('%Y-%m-%d') if not isinstance(record['account_date'], str) else record['account_date']
            owner = record['owner'] or '未知'
            record_type = record['record_type']
            total_amount = float(record['total_amount'])
            
            if date_str in calendar_data:
                if owner not in calendar_data[date_str]['owners']:
                    calendar_data[date_str]['owners'][owner] = {
                        'income': 0.0,
                        'expense': 0.0,
                        'total': 0.0
                    }
                
                if record_type == '收入':
                    calendar_data[date_str]['owners'][owner]['income'] = total_amount
                else:
                    calendar_data[date_str]['owners'][owner]['expense'] = total_amount
                
                calendar_data[date_str]['owners'][owner]['total'] = (
                    calendar_data[date_str]['owners'][owner]['income'] - 
                    calendar_data[date_str]['owners'][owner]['expense']
                )
        
        # 计算每天的汇总
        for date_str, day_data in calendar_data.items():
            day_income = 0.0
            day_expense = 0.0
            
            for owner_data in day_data['owners'].values():
                day_income += owner_data['income']
                day_expense += owner_data['expense']
            
            day_data['summary'] = {
                'total_income': day_income,
                'total_expense': day_expense,
                'net_amount': day_income - day_expense,
                'owner_count': len(day_data['owners'])
            }
        
        cursor.close()
        
        # 获取所有所属人（用于前端显示）
        owners_query = "SELECT DISTINCT owner FROM daily_accounts WHERE owner IS NOT NULL AND owner != ''"
        cursor = connection.cursor(dictionary=True)
        cursor.execute(owners_query)
        owners_result = cursor.fetchall()
        all_owners = [owner['owner'] for owner in owners_result]
        cursor.close()
        
        return jsonify({
            'success': True,
            'year': year,
            'month': month,
            'owners': all_owners,
            'selected_owner': owner,
            'calendar_data': list(calendar_data.values()),
            'month_name': f"{year}年{month:02d}月"
        })
        
    except Error as e:
        logger.error(f"获取日历数据错误: {e}")
        return jsonify({'success': False, 'message': '获取日历数据失败'}), 500
    except Exception as e:
        logger.error(f"获取日历数据异常: {str(e)}")
        return jsonify({'success': False, 'message': '获取日历数据失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 修复：详细记账统计API =====================
@app.route('/api/account/statistics/detailed')
@login_required
def get_detailed_account_statistics():
    """获取详细的记账统计信息"""
    try:
        stat_type = request.args.get('type', 'monthly')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner', '全部')
        
        # 如果没有提供日期范围，默认使用最近一年
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        statistics = get_account_statistics_by_period(stat_type, start_date, end_date, owner)
        summary = get_account_summary_statistics(start_date, end_date, owner)
        
        return jsonify({
            'statistics': statistics,
            'summary': summary,
            'filters': {
                'type': stat_type,
                'start_date': start_date,
                'end_date': end_date,
                'owner': owner
            }
        })
    except Exception as e:
        logger.error(f"获取详细记账统计错误: {str(e)}")
        return jsonify({
            'statistics': [],
            'summary': {
                'total': {'total_count': 0, 'total_expense': 0.0, 'total_income': 0.0},
                'by_owner': [],
                'by_category': []
            },
            'filters': {}
        })
    
    
        
@app.route('/api/account/records/search', methods=['POST'])
@login_required
def search_account_records():
    """搜索记账记录（带分页）- 支持开始时间和结束时间"""
    data = request.json
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 获取分页参数
        page = data.get('page', 1)
        per_page = data.get('per_page', 20)
        offset = (page - 1) * per_page
        
        # 构建查询条件
        query = """
            SELECT id, record_type, category, subcategory, amount, account_date, 
                   description, payment_method, owner
            FROM daily_accounts 
            WHERE 1=1
        """
        count_query = "SELECT COUNT(*) as total FROM daily_accounts WHERE 1=1"
        params = []

        # 记录类型筛选
        record_type_filter = data.get('record_type', '全部')
        if record_type_filter != '全部':
            query += " AND record_type = %s"
            count_query += " AND record_type = %s"
            params.append(record_type_filter)

        # 类别筛选
        category_filter = data.get('category', '全部')
        if category_filter != '全部':
            query += " AND category = %s"
            count_query += " AND category = %s"
            params.append(category_filter)

        # 子类别筛选
        subcategory_filter = data.get('subcategory', '全部')
        if subcategory_filter != '全部':
            query += " AND subcategory = %s"
            count_query += " AND subcategory = %s"
            params.append(subcategory_filter)

        # 日期范围筛选（新增开始时间和结束时间）
        start_date = data.get('start_date', '').strip()
        end_date = data.get('end_date', '').strip()
        
        if start_date and end_date:
            query += " AND account_date BETWEEN %s AND %s"
            count_query += " AND account_date BETWEEN %s AND %s"
            params.extend([start_date, end_date])
        elif start_date:
            query += " AND account_date >= %s"
            count_query += " AND account_date >= %s"
            params.append(start_date)
        elif end_date:
            query += " AND account_date <= %s"
            count_query += " AND account_date <= %s"
            params.append(end_date)

        # 所属人筛选
        owner_filter = data.get('owner', '全部')
        if owner_filter != '全部':
            query += " AND owner = %s"
            count_query += " AND owner = %s"
            params.append(owner_filter)

        # 获取总数
        cursor.execute(count_query, params)
        total_result = cursor.fetchone()
        total_count = total_result['total'] if total_result else 0

        # 添加排序和分页
        query += " ORDER BY account_date DESC, id DESC LIMIT %s OFFSET %s"
        params.extend([per_page, offset])

        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # 计算统计信息
        stats_query = """
            SELECT 
                COUNT(*) as total_count,
                COALESCE(SUM(CASE WHEN record_type = '支出' THEN amount ELSE 0 END), 0) as total_expense,
                COALESCE(SUM(CASE WHEN record_type = '收入' THEN amount ELSE 0 END), 0) as total_income
            FROM daily_accounts 
            WHERE 1=1
        """
        stats_params = params[:-2] if len(params) > 2 else []  # 移除分页参数
        
        # 构建统计查询的WHERE条件
        where_index = 0
        if 'record_type' in data and data['record_type'] != '全部':
            stats_query += " AND record_type = %s"
        if 'category' in data and data['category'] != '全部':
            stats_query += " AND category = %s"
        if 'subcategory' in data and data['subcategory'] != '全部':
            stats_query += " AND subcategory = %s"
        if start_date and end_date:
            stats_query += " AND account_date BETWEEN %s AND %s"
        elif start_date:
            stats_query += " AND account_date >= %s"
        elif end_date:
            stats_query += " AND account_date <= %s"
        if 'owner' in data and data['owner'] != '全部':
            stats_query += " AND owner = %s"
        
        cursor.execute(stats_query, stats_params)
        stats_result = cursor.fetchone()
        
        for record in records:
            record['id'] = int(record['id'])
            record['amount'] = float(record['amount'])
            if record['account_date'] and not isinstance(record['account_date'], str):
                record['account_date'] = record['account_date'].strftime("%Y-%m-%d")
            if record['subcategory'] is None:
                record['subcategory'] = ""
            if record['description'] is None:
                record['description'] = ""
            if record['payment_method'] is None:
                record['payment_method'] = "现金"
            if record['owner'] is None:
                record['owner'] = "郭宁"

        cursor.close()
        
        return jsonify({
            'records': records,
            'pagination': {
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': (total_count + per_page - 1) // per_page
            },
            'stats': {
                'total_count': stats_result['total_count'] if stats_result else 0,
                'total_expense': float(stats_result['total_expense']) if stats_result else 0.0,
                'total_income': float(stats_result['total_income']) if stats_result else 0.0,
                'net_amount': float(stats_result['total_income']) - float(stats_result['total_expense']) if stats_result else 0.0
            }
        })
        
    except Error as e:
        logger.error(f"搜索记账记录错误: {e}")
        return jsonify({'error': '搜索记录失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：Excel导入功能 =====================
@app.route('/api/account/import', methods=['POST'])
@login_required
def import_account_data():
    """导入Excel记账数据"""
    try:
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '只支持Excel文件(.xlsx, .xls)'})
        
        # 读取Excel文件
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active
        
        # 获取表头 - 修复：更灵活的表头识别
        headers = []
        for cell in ws[1]:
            header_value = str(cell.value).strip() if cell.value else ""
            headers.append(header_value)
        
        logger.info(f"读取到的表头: {headers}")
        
        # 创建表头映射 - 支持多种可能的表头名称
        header_mapping = {
            '记录类型': ['记录类型', '类型', '收支类型', 'record_type'],
            '类别': ['类别', '分类', 'category'],
            '子类别': ['子类别', '子分类', 'subcategory'],
            '金额': ['金额', '数额', 'money', 'amount'],
            '日期': ['日期', '时间', 'date', 'account_date'],
            '描述': ['描述', '备注', '说明', 'description', 'remark'],
            '支付方式': ['支付方式', '付款方式', '支付方法', 'payment_method'],
            '所属人': ['所属人', '所有人', '负责人', 'owner']
        }
        
        # 构建实际表头到标准表头的映射
        actual_to_standard = {}
        for standard_header, possible_headers in header_mapping.items():
            for actual_header in headers:
                if actual_header in possible_headers:
                    actual_to_standard[actual_header] = standard_header
                    break
        
        # 检查必要字段
        required_fields = ['记录类型', '类别', '金额', '日期']
        missing_fields = []
        for field in required_fields:
            if field not in actual_to_standard.values():
                missing_fields.append(field)
        
        if missing_fields:
            return jsonify({
                'success': False, 
                'message': f'Excel文件缺少必要列: {", ".join(missing_fields)}。请确保包含以下列: {", ".join(required_fields)}。实际表头: {", ".join(headers)}'
            })
        
        # 处理数据
        imported_count = 0
        duplicate_count = 0  # 添加重复计数初始化
        error_count = 0
        error_messages = []
        duplicate_messages = []  # 添加重复消息列表初始化
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not any(row):
                continue
            
            # 创建数据字典 - 使用标准表头
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i] in actual_to_standard:
                    standard_header = actual_to_standard[headers[i]]
                    row_data[standard_header] = value
            
            try:
                # 验证必要字段
                if not row_data.get('记录类型') or not row_data.get('类别') or not row_data.get('金额') or not row_data.get('日期'):
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 缺少必要字段")
                    continue
                
                # 验证金额
                try:
                    amount = float(row_data['金额'])
                    if amount <= 0:
                        error_count += 1
                        error_messages.append(f"第{row_num}行: 金额必须大于0")
                        continue
                except (ValueError, TypeError):
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 金额格式错误 - {row_data['金额']}")
                    continue
                
                # 验证日期 - 修复：支持更多日期格式
                try:
                    account_date = None
                    date_value = row_data['日期']
                    
                    if isinstance(date_value, datetime):
                        account_date = date_value.strftime('%Y-%m-%d')
                    elif isinstance(date_value, str):
                        # 尝试多种日期格式
                        date_formats = [
                            '%Y-%m-%d',     # 2024-01-01
                            '%Y/%m/%d',     # 2024/01/01
                            '%Y.%m.%d',     # 2024.01.01
                            '%Y年%m月%d日',  # 2024年01月01日
                        ]
                        
                        for date_format in date_formats:
                            try:
                                parsed_date = datetime.strptime(date_value, date_format)
                                account_date = parsed_date.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                        
                        if not account_date:
                            # 如果以上格式都不匹配，尝试使用dateutil解析（如果可用）
                            try:
                                from dateutil import parser
                                parsed_date = parser.parse(date_value)
                                account_date = parsed_date.strftime('%Y-%m-%d')
                            except:
                                raise ValueError("无法解析日期")
                    else:
                        error_count += 1
                        error_messages.append(f"第{row_num}行: 日期格式错误 - {date_value}")
                        continue
                        
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 日期格式错误 - {row_data['日期']}")
                    continue
                
                # 验证记录类型
                record_type = str(row_data['记录类型']).strip()
                if record_type not in ['支出', '收入']:
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 记录类型必须是'支出'或'收入'")
                    continue
                
                # 构建记录数据
                record = {
                    'record_type': record_type,
                    'owner': str(row_data.get('所属人', '郭宁')).strip(),
                    'category': str(row_data.get('类别', '')).strip(),
                    'subcategory': str(row_data.get('子类别', '')).strip(),
                    'amount': amount,
                    'account_date': account_date,
                    'description': str(row_data.get('描述', '')).strip(),
                    'payment_method': str(row_data.get('支付方式', '现金')).strip()
                }
                
                # 检查是否重复
                if is_duplicate_account_record(record):
                    duplicate_count += 1
                    duplicate_messages.append(f"第{row_num}行: 记录已存在，跳过导入")
                    continue
                
                # 保存记录
                result = save_account_record(record)
                if result == 'duplicate':
                    # 这种情况不应该发生，因为我们已经检查过了，但为了安全还是处理
                    duplicate_count += 1
                    duplicate_messages.append(f"第{row_num}行: 记录已存在，跳过导入")
                elif result:
                    imported_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 保存到数据库失败")
                    
            except Exception as e:
                error_count += 1
                error_messages.append(f"第{row_num}行: 处理错误 - {str(e)}")
                logger.error(f"处理第{row_num}行时出错: {e}")
                continue
        
        # 记录操作日志
        log_operation("IMPORT", 
                     f"导入记账数据 - 成功: {imported_count}条, 重复: {duplicate_count}条, 失败: {error_count}条", 
                     user_name=session.get('username', 'admin'))
        
        result = {
            'success': True,
            'message': f'导入完成！成功导入 {imported_count} 条记录，跳过 {duplicate_count} 条重复记录，失败 {error_count} 条记录。',
            'imported_count': imported_count,
            'duplicate_count': duplicate_count,
            'error_count': error_count
        }
        
        # 如果有错误或重复，添加详细信息
        if error_messages:
            result['error_messages'] = error_messages[:10]  # 只返回前10条错误信息
        if duplicate_messages:
            result['duplicate_messages'] = duplicate_messages[:10]  # 只返回前10条重复信息
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"导入Excel数据错误: {str(e)}")
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})
        
       
#模板下载函数
@app.route('/api/account/template')
@login_required
def download_account_template():
    """下载Excel导入模板"""
    try:
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "记账数据导入模板"
        
        # 设置表头样式
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # 表头 - 使用标准名称
        headers = [
            '记录类型', '类别', '子类别', '金额', '日期', 
            '描述', '支付方式', '所属人'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        example_data = [
            ['支出', '食品酒水', '午餐', 20.00, '2024-01-01', '午餐费用', '支付宝', '郭宁'],
            ['支出', '衣服饰品', '裤子', 200.00, '2024-01-02', '', '微信', '李佳慧'],
            ['收入', '工资收入', '工资', 5000.00, '2024-01-03', '本月工资', '银行卡', '郭宁']
        ]
        
        for row_idx, data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # 添加说明
        ws.cell(row=5, column=1, value="说明:").font = Font(bold=True)
        ws.cell(row=6, column=1, value="1. 所有列都是必填项，只有描述列可为空")
        ws.cell(row=7, column=1, value="2. 记录类型: 支出 或 收入")
        ws.cell(row=8, column=1, value="3. 日期格式: YYYY-MM-DD 或 YYYY/MM/DD")
        ws.cell(row=9, column=1, value="4. 金额: 必须大于0的数字")
        ws.cell(row=10, column=1, value="5. 所属人: 郭宁 或 李佳慧")
        
        # 设置列宽
        column_widths = [12, 15, 15, 12, 12, 20, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 创建内存文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="记账数据导入模板.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"下载模板错误: {str(e)}")
        return jsonify({'error': '下载模板失败'}), 500


# ===================== 新增：礼尚往来记录Excel导入功能 =====================
@app.route('/api/gift_records/import', methods=['POST'])
@login_required
def import_gift_records():
    """导入Excel礼尚往来记录数据"""
    try:
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '只支持Excel文件(.xlsx, .xls)'})
        
        # 读取Excel文件
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active
        
        # 获取表头 - 清理表头中的特殊字符
        headers = []
        for cell in ws[1]:
            header_value = str(cell.value).strip() if cell.value else ""
            # 清理表头：移除星号等特殊字符
            cleaned_header = header_value.replace('*', '').strip()
            headers.append(cleaned_header)
        
        logger.info(f"读取到的礼尚往来记录表头: {headers}")
        
        # 直接检查清理后的表头是否包含必要字段
        required_fields = ['记录类型', '姓名', '金额', '事件', '日期']
        missing_fields = []
        for field in required_fields:
            if field not in headers:
                missing_fields.append(field)
        
        if missing_fields:
            return jsonify({
                'success': False, 
                'message': f'Excel文件缺少必要列: {", ".join(missing_fields)}。请确保包含以下列: {", ".join(required_fields)}。实际表头: {", ".join(headers)}'
            })
        
        # 处理数据
        imported_count = 0
        duplicate_count = 0  # 新增：重复记录计数
        error_count = 0
        error_messages = []
        duplicate_messages = []  # 新增：重复记录消息
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not any(row):
                continue
            
            # 创建数据字典
            row_data = dict(zip(headers, row))
            
            try:
                # 验证必要字段
                if (not row_data.get('记录类型') or not row_data.get('姓名') or 
                    not row_data.get('金额') or not row_data.get('事件') or not row_data.get('日期')):
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 缺少必要字段")
                    continue
                
                # 验证金额
                try:
                    amount = float(row_data['金额'])
                    if amount <= 0:
                        error_count += 1
                        error_messages.append(f"第{row_num}行: 金额必须大于0")
                        continue
                except (ValueError, TypeError):
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 金额格式错误 - {row_data['金额']}")
                    continue
                
                # 验证回礼金额
                return_amount = 0.0
                if row_data.get('回礼金额'):
                    try:
                        return_amount = float(row_data['回礼金额'])
                        if return_amount < 0:
                            error_count += 1
                            error_messages.append(f"第{row_num}行: 回礼金额不能为负数")
                            continue
                    except (ValueError, TypeError):
                        return_amount = 0.0
                
                # 验证日期
                try:
                    date_value = row_data['日期']
                    formatted_date = None
                    
                    if isinstance(date_value, datetime):
                        formatted_date = date_value.strftime('%Y-%m-%d')
                    elif isinstance(date_value, str):
                        # 尝试多种日期格式
                        date_formats = [
                            '%Y-%m-%d',     # 2024-01-01
                            '%Y/%m/%d',     # 2024/01/01
                            '%Y.%m.%d',     # 2024.01.01
                            '%Y年%m月%d日',  # 2024年01月01日
                        ]
                        
                        for date_format in date_formats:
                            try:
                                parsed_date = datetime.strptime(date_value, date_format)
                                formatted_date = parsed_date.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                        
                        if not formatted_date:
                            # 如果以上格式都不匹配，尝试使用dateutil解析（如果可用）
                            try:
                                from dateutil import parser
                                parsed_date = parser.parse(date_value)
                                formatted_date = parsed_date.strftime('%Y-%m-%d')
                            except:
                                raise ValueError("无法解析日期")
                    else:
                        error_count += 1
                        error_messages.append(f"第{row_num}行: 日期格式错误 - {date_value}")
                        continue
                        
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 日期格式错误 - {row_data['日期']}")
                    continue
                
                # 验证回礼日期
                return_date = None
                if row_data.get('回礼日期'):
                    try:
                        return_date_value = row_data['回礼日期']
                        if isinstance(return_date_value, datetime):
                            return_date = return_date_value.strftime('%Y-%m-%d')
                        elif isinstance(return_date_value, str):
                            # 尝试多种日期格式
                            for date_format in date_formats:
                                try:
                                    parsed_date = datetime.strptime(return_date_value, date_format)
                                    return_date = parsed_date.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    continue
                            
                            if not return_date:
                                try:
                                    from dateutil import parser
                                    parsed_date = parser.parse(return_date_value)
                                    return_date = parsed_date.strftime('%Y-%m-%d')
                                except:
                                    return_date = None
                    except Exception as e:
                        # 回礼日期不是必须的，如果有错误可以忽略
                        return_date = None
                
                # 验证记录类型
                record_type = str(row_data['记录类型']).strip()
                if record_type not in ['受礼记录', '随礼记录']:
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 记录类型必须是'受礼记录'或'随礼记录'")
                    continue
                
                # 构建记录数据
                record = {
                    'record_type': record_type,
                    'owner': str(row_data.get('所属人', '郭宁')).strip(),
                    'name': str(row_data.get('姓名', '')).strip(),
                    'amount': amount,
                    'occasion': str(row_data.get('事件', '')).strip(),
                    'date': formatted_date,
                    'return_amount': return_amount,
                    'return_occasion': str(row_data.get('回礼事件', '')).strip(),
                    'return_date': return_date,
                    'remark': str(row_data.get('备注', '')).strip()
                }
                
                # ==================== 新增：检查重复记录 ====================
                if is_duplicate_gift_record(record):
                    duplicate_count += 1
                    duplicate_messages.append(f"第{row_num}行: 记录已存在，跳过导入")
                    continue
                # ==================== 结束新增 ====================
                
                # 保存记录
                if save_record(record):
                    imported_count += 1
                else:
                    error_count += 1
                    error_messages.append(f"第{row_num}行: 保存到数据库失败")
                    
            except Exception as e:
                error_count += 1
                error_messages.append(f"第{row_num}行: 处理错误 - {str(e)}")
                logger.error(f"处理第{row_num}行时出错: {e}")
                continue
        
        # 记录操作日志
        log_operation("IMPORT", 
                     f"导入礼尚往来记录 - 成功: {imported_count}条, 重复: {duplicate_count}条, 失败: {error_count}条", 
                     user_name=session.get('username', 'admin'))
        
        result = {
            'success': True,
            'message': f'导入完成！成功导入 {imported_count} 条记录，跳过 {duplicate_count} 条重复记录，失败 {error_count} 条记录。',
            'imported_count': imported_count,
            'duplicate_count': duplicate_count,  # 新增：返回重复计数
            'error_count': error_count
        }
        
        # 如果有错误或重复，添加详细信息
        if error_messages:
            result['error_messages'] = error_messages[:10]  # 只返回前10条错误信息
        if duplicate_messages:
            result['duplicate_messages'] = duplicate_messages[:10]  # 只返回前10条重复信息
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"导入礼尚往来记录Excel数据错误: {str(e)}")
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

@app.route('/api/gift_records/template')
@login_required
def download_gift_records_template():
    """下载礼尚往来记录Excel导入模板"""
    try:
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "礼尚往来记录导入模板"
        
        # 设置表头样式
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # 表头
        headers = [
            '记录类型', '姓名', '金额', '事件', '日期', 
            '回礼金额', '回礼事件', '回礼日期', '备注', '所属人'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        example_data = [
            ['受礼记录', '张三', 500.00, '结婚礼金', '2024-01-01', 0, '', '', '同事结婚', '郭宁'],
            ['随礼记录', '李四', 300.00, '生日礼物', '2024-01-02', 200, '回礼', '2024-02-01', '', '李佳慧'],
            ['受礼记录', '王五', 1000.00, '节日红包', '2024-01-03', 800, '回礼红包', '2024-02-02', '春节红包', '郭宁']
        ]
        
        for row_idx, data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # 添加说明
        ws.cell(row=5, column=1, value="说明:").font = Font(bold=True)
        ws.cell(row=6, column=1, value="1. 以下列为必填项: 记录类型, 姓名, 金额, 事件, 日期")
        ws.cell(row=7, column=1, value="2. 记录类型: 受礼记录 或 随礼记录")
        ws.cell(row=8, column=1, value="3. 日期格式: YYYY-MM-DD 或 YYYY/MM/DD")
        ws.cell(row=9, column=1, value="4. 金额: 必须大于0的数字")
        ws.cell(row=10, column=1, value="5. 回礼金额: 可选，如果填写必须大于等于0")
        ws.cell(row=11, column=1, value="6. 所属人: 郭宁 或 李佳慧")
        
        # 设置列宽
        column_widths = [12, 12, 12, 20, 12, 12, 15, 12, 20, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 创建内存文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="礼尚往来记录导入模板.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"下载礼尚往来记录模板错误: {str(e)}")
        return jsonify({'error': '下载模板失败'}), 500


@app.route('/api/account/export')
@login_required
def export_account_data():
    """导出记账数据到Excel - 修复版，支持日期范围"""
    try:
        # 获取查询参数
        record_type = request.args.get('record_type', '全部')
        category = request.args.get('category', '全部')
        subcategory = request.args.get('subcategory', '全部')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        owner = request.args.get('owner', '全部')
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件 - 修复：支持日期范围
        query = """
            SELECT record_type, category, subcategory, amount, account_date, 
                   description, payment_method, owner
            FROM daily_accounts 
            WHERE 1=1
        """
        params = []

        # 记录类型筛选
        if record_type != '全部':
            query += " AND record_type = %s"
            params.append(record_type)

        # 类别筛选
        if category != '全部':
            query += " AND category = %s"
            params.append(category)

        # 子类别筛选
        if subcategory != '全部':
            query += " AND subcategory = %s"
            params.append(subcategory)

        # 日期范围筛选 - 修复：正确处理日期范围
        if start_date and end_date:
            query += " AND account_date BETWEEN %s AND %s"
            params.extend([start_date, end_date])
        elif start_date:
            query += " AND account_date >= %s"
            params.append(start_date)
        elif end_date:
            query += " AND account_date <= %s"
            params.append(end_date)

        # 所属人筛选
        if owner != '全部':
            query += " AND owner = %s"
            params.append(owner)

        query += " ORDER BY account_date DESC, id DESC"

        logger.info(f"导出查询SQL: {query}")
        logger.info(f"导出查询参数: {params}")

        cursor.execute(query, params)
        records = cursor.fetchall()
        cursor.close()

        logger.info(f"导出记录数: {len(records)}")

        # 如果没有记录，返回提示
        if not records:
            return jsonify({
                'success': False,
                'message': f'没有找到符合条件的记录。查询条件: 开始日期={start_date}, 结束日期={end_date}, 记录类型={record_type}, 类别={category}, 子类别={subcategory}, 所属人={owner}'
            })

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "记账数据"

        # 设置标题样式
        title_font = Font(size=16, bold=True)
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 写入标题
        title = "记账数据导出报告"
        if start_date or end_date:
            title += f" ({start_date} 至 {end_date})"
        
        ws['A1'] = title
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')

        # 写入筛选条件
        filter_text = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        if start_date or end_date:
            filter_text += f" | 日期范围: {start_date if start_date else '不限'} 至 {end_date if end_date else '不限'}"
        if record_type != '全部':
            filter_text += f" | 记录类型: {record_type}"
        if category != '全部':
            filter_text += f" | 类别: {category}"
        if subcategory != '全部':
            filter_text += f" | 子类别: {subcategory}"
        if owner != '全部':
            filter_text += f" | 所属人: {owner}"
        
        ws['A2'] = filter_text
        ws.merge_cells('A2:H2')

        # 写入汇总信息
        total_count = len(records)
        total_expense = sum(float(r['amount']) for r in records if r['record_type'] == '支出')
        total_income = sum(float(r['amount']) for r in records if r['record_type'] == '收入')
        
        ws['A3'] = f"汇总: 共 {total_count} 条记录，总支出 ¥{total_expense:.2f}，总收入 ¥{total_income:.2f}，净收入 ¥{total_income - total_expense:.2f}"
        ws.merge_cells('A3:H3')

        # 写入表头
        headers = ['记录类型', '类别', '子类别', '金额', '日期', '描述', '支付方式', '所属人']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row, record in enumerate(records, 6):
            ws.cell(row=row, column=1, value=record['record_type']).border = border
            ws.cell(row=row, column=2, value=record['category']).border = border
            ws.cell(row=row, column=3, value=record['subcategory'] or '').border = border
            ws.cell(row=row, column=4, value=float(record['amount'])).border = border
            ws.cell(row=row, column=5, value=record['account_date'].strftime('%Y-%m-%d') if not isinstance(record['account_date'], str) else record['account_date']).border = border
            ws.cell(row=row, column=6, value=record['description'] or '').border = border
            ws.cell(row=row, column=7, value=record['payment_method'] or '现金').border = border
            ws.cell(row=row, column=8, value=record['owner'] or '郭宁').border = border

            # 为支出记录设置红色字体
            if record['record_type'] == '支出':
                ws.cell(row=row, column=4).font = Font(color="FF0000")

        # 设置列宽
        column_widths = [12, 15, 15, 12, 12, 30, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 设置金额列的格式
        for row in range(6, len(records) + 6):
            ws.cell(row=row, column=4).number_format = '0.00'

        # 添加自动筛选
        ws.auto_filter.ref = f"A5:H{5 + len(records)}"

        # 创建内存文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        filename = f"记账数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        if start_date or end_date:
            filename += f"_{start_date if start_date else '起始'}_至_{end_date if end_date else '结束'}"
        filename += ".xlsx"

        logger.info(f"生成导出文件: {filename}，大小: {output.getbuffer().nbytes} 字节")

        # 记录操作日志
        log_operation("EXPORT", 
                     f"导出记账数据 - 记录数: {len(records)}, 日期范围: {start_date} 至 {end_date}",
                     user_name=session.get('username', 'admin'))

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"导出记账数据错误: {str(e)}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：记账统计导出路由 =====================
@app.route('/api/account/statistics/export')
@login_required
def export_account_statistics():
    """导出记账统计信息"""
    try:
        stat_type = request.args.get('type', 'monthly')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner', '全部')
        
        # 获取统计数据和汇总
        statistics = get_account_statistics_by_period(stat_type, start_date, end_date, owner)
        summary = get_account_summary_statistics(start_date, end_date, owner)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "记账统计"
        
        # 设置标题
        title_font = Font(size=16, bold=True)
        header_font = Font(bold=True)
        
        # 写入标题
        ws['A1'] = f"记账统计报告 - {start_date} 至 {end_date}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        # 写入汇总信息
        ws['A3'] = "汇总统计"
        ws['A3'].font = header_font
        
        total = summary.get('total', {})
        ws['A4'] = "总记录数"
        ws['B4'] = total.get('total_count', 0)
        
        ws['A5'] = "总支出"
        ws['B5'] = float(total.get('total_expense', 0))
        
        ws['A6'] = "总收入"
        ws['B6'] = float(total.get('total_income', 0))
        
        ws['A7'] = "净收入"
        ws['B7'] = float(total.get('total_income', 0)) - float(total.get('total_expense', 0))
        
        # 写入详细统计
        row = 9
        if stat_type == 'monthly' and statistics:
            ws[f'A{row}'] = "月度统计"
            ws[f'A{row}'].font = header_font
            row += 1
            
            headers = ['年月', '记录类型', '所属人', '记录数', '总金额']
            for i, header in enumerate(headers):
                ws.cell(row=row, column=i+1, value=header).font = header_font
            
            row += 1
            for stat in statistics:
                month_name = stat.get('month_name', f"{stat.get('year', '')}年{stat.get('month', '')}月")
                ws.cell(row=row, column=1, value=month_name)
                ws.cell(row=row, column=2, value=stat.get('record_type', ''))
                ws.cell(row=row, column=3, value=stat.get('owner', ''))
                ws.cell(row=row, column=4, value=stat.get('count', 0))
                ws.cell(row=row, column=5, value=float(stat.get('total_amount', 0)))
                row += 1
        
        # 创建内存文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"记账统计_{start_date}_至_{end_date}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"导出记账统计错误: {str(e)}")
        return jsonify({'error': '导出统计信息失败'}), 500


# ===================== 新增：图表数据API =====================
# ===================== 修复：图表数据API - 修正所属人收支对比数据 =====================
@app.route('/api/account/statistics/charts')
@login_required
def get_account_charts_data():
    """获取图表数据"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner', '全部')
        
        # 如果没有提供日期范围，默认使用最近一年
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        logger.info(f"获取图表数据 - 日期: {start_date} 到 {end_date}, 所属人: {owner}")
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        try:
            cursor = connection.cursor(dictionary=True)
            
            # 构建查询条件
            where_conditions = ["account_date BETWEEN %s AND %s"]
            params = [start_date, end_date]
            
            if owner and owner != "全部":
                where_conditions.append("owner = %s")
                params.append(owner)
            
            where_clause = " AND ".join(where_conditions)
            
            # 月度趋势数据
            monthly_query = f"""
                SELECT 
                    YEAR(account_date) as year,
                    MONTH(account_date) as month,
                    record_type,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                WHERE {where_clause}
                GROUP BY YEAR(account_date), MONTH(account_date), record_type
                ORDER BY year, month
            """
            
            cursor.execute(monthly_query, params)
            monthly_data = cursor.fetchall()
            
            # 处理月度数据
            months = []
            income_by_month = {}
            expense_by_month = {}
            
            for record in monthly_data:
                month_key = f"{record['year']}-{record['month']:02d}"
                if month_key not in months:
                    months.append(month_key)
                
                if record['record_type'] == '收入':
                    income_by_month[month_key] = float(record['total_amount'])
                else:
                    expense_by_month[month_key] = float(record['total_amount'])
            
            # 填充缺失的月份数据
            monthly_income = [income_by_month.get(month, 0) for month in months]
            monthly_expense = [expense_by_month.get(month, 0) for month in months]
            
            # 获取所属人收支数据 - 修正：直接从数据库查询实际数据
            owners_query = f"""
                SELECT 
                    owner,
                    record_type,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                WHERE {where_clause}
                GROUP BY owner, record_type
                HAVING owner IN ('郭宁', '李佳慧')
                ORDER BY owner, record_type
            """
            
            cursor.execute(owners_query, params)
            owners_data = cursor.fetchall()
            
            # 处理所属人数据
            owner_income = {'郭宁': 0, '李佳慧': 0}
            owner_expense = {'郭宁': 0, '李佳慧': 0}
            
            for record in owners_data:
                if record['record_type'] == '收入':
                    owner_income[record['owner']] = float(record['total_amount'])
                else:
                    owner_expense[record['owner']] = float(record['total_amount'])
            
            logger.info(f"所属人收支数据 - 郭宁: 收入={owner_income['郭宁']}, 支出={owner_expense['郭宁']}")
            logger.info(f"所属人收支数据 - 李佳慧: 收入={owner_income['李佳慧']}, 支出={owner_expense['李佳慧']}")
            
            # 获取类别统计数据（用于饼图）
            expense_category_query = f"""
                SELECT 
                    category,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                WHERE {where_clause} AND record_type = '支出'
                GROUP BY category
                ORDER BY total_amount DESC
                LIMIT 8
            """
            
            income_category_query = f"""
                SELECT 
                    category,
                    COALESCE(SUM(amount), 0) as total_amount
                FROM daily_accounts 
                WHERE {where_clause} AND record_type = '收入'
                GROUP BY category
                ORDER BY total_amount DESC
                LIMIT 5
            """
            
            cursor.execute(expense_category_query, params)
            expense_categories_data = cursor.fetchall()
            
            cursor.execute(income_category_query, params)
            income_categories_data = cursor.fetchall()
            
            # 处理类别数据
            expense_categories = [item['category'] for item in expense_categories_data]
            expense_category_amounts = [float(item['total_amount']) for item in expense_categories_data]
            
            income_categories = [item['category'] for item in income_categories_data]
            income_category_amounts = [float(item['total_amount']) for item in income_categories_data]
            
            # 构建返回数据
            chart_data = {
                'monthly': {
                    'labels': months,
                    'income': monthly_income,
                    'expense': monthly_expense
                },
                'quarterly': {
                    'labels': ['Q1', 'Q2', 'Q3', 'Q4'],
                    'income': [sum(monthly_income[0:3]), sum(monthly_income[3:6]), sum(monthly_income[6:9]), sum(monthly_income[9:12])],
                    'expense': [sum(monthly_expense[0:3]), sum(monthly_expense[3:6]), sum(monthly_expense[6:9]), sum(monthly_expense[9:12])]
                },
                'yearly': {
                    'labels': list(set([int(month.split('-')[0]) for month in months])),
                    'income': [],
                    'expense': []
                },
                'category': {
                    'expense': {
                        'labels': expense_categories,
                        'data': expense_category_amounts
                    },
                    'income': {
                        'labels': income_categories,
                        'data': income_category_amounts
                    }
                },
                'comparison': {
                    'balance': {
                        'labels': months,
                        'income': monthly_income,
                        'expense': monthly_expense,
                        'net': [income - expense for income, expense in zip(monthly_income, monthly_expense)]
                    },
                    'owners': {
                        'labels': ['郭宁', '李佳慧'],
                        'income': [owner_income['郭宁'], owner_income['李佳慧']],
                        'expense': [owner_expense['郭宁'], owner_expense['李佳慧']]
                    }
                }
            }
            
            # 处理年度数据
            yearly_labels = chart_data['yearly']['labels']
            yearly_labels.sort()
            yearly_income = []
            yearly_expense = []
            
            for year in yearly_labels:
                year_income = 0
                year_expense = 0
                for month in months:
                    if month.startswith(str(year)):
                        month_index = months.index(month)
                        year_income += monthly_income[month_index]
                        year_expense += monthly_expense[month_index]
                yearly_income.append(year_income)
                yearly_expense.append(year_expense)
            
            chart_data['yearly']['income'] = yearly_income
            chart_data['yearly']['expense'] = yearly_expense
            
            # 将年度标签转换为字符串
            chart_data['yearly']['labels'] = [f"{year}年" for year in yearly_labels]
            
            cursor.close()
            return jsonify(chart_data)
            
        except Error as e:
            logger.error(f"获取图表数据错误: {e}")
            return jsonify({'error': '获取图表数据失败'}), 500
        finally:
            if connection and connection.is_connected():
                connection.close()
                
    except Exception as e:
        logger.error(f"获取图表数据异常: {str(e)}")
        return jsonify({'error': '获取图表数据失败'}), 500



# ===================== 修复：简化类别统计图表API =====================
# ===================== 修复：完整类别统计图表API =====================
@app.route('/api/account/statistics/categories')
@login_required
def get_category_charts_data():
    """获取类别统计图表数据（完整版）"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        owner = request.args.get('owner', '全部')
        time_range = request.args.get('time_range', 'all')
        
        # 如果没有提供日期范围，默认使用最近一年
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        logger.info(f"获取类别统计图表数据 - 时间范围: {time_range}, 日期: {start_date} 到 {end_date}, 所属人: {owner}")
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        try:
            cursor = connection.cursor(dictionary=True)
            
            # 构建查询条件
            where_conditions = ["account_date BETWEEN %s AND %s"]
            params = [start_date, end_date]
            
            if owner and owner != "全部":
                where_conditions.append("owner = %s")
                params.append(owner)
            
            where_clause = " AND ".join(where_conditions)
            
            if time_range == 'all':
                # 全部数据 - 简单的类别统计
                logger.info(f"执行简单类别统计查询")
                result = get_simple_category_data(cursor, where_clause, params)
            else:
                # 按时间维度统计
                logger.info(f"执行时间维度类别统计查询 - 时间范围: {time_range}")
                result = get_time_based_category_data(cursor, where_clause, params, time_range)
            
            # 记录结果结构
            logger.info(f"类别统计结果结构 - 包含expense: {'expense' in result}, 包含income: {'income' in result}")
            if 'expense' in result:
                logger.info(f"支出数据结构 - labels数量: {len(result['expense'].get('labels', []))}, datasets数量: {len(result['expense'].get('datasets', []))}")
            if 'income' in result:
                logger.info(f"收入数据结构 - labels数量: {len(result['income'].get('labels', []))}, datasets数量: {len(result['income'].get('datasets', []))}")
            
            # 确保数据格式统一
            result = format_category_data(result)
            
            # 记录格式化后的结构
            logger.info(f"格式化后 - 支出datasets数量: {len(result['expense'].get('datasets', []))}, 收入datasets数量: {len(result['income'].get('datasets', []))}")
            
            cursor.close()
            return jsonify(result)
                
        except Error as e:
            logger.error(f"获取类别统计图表数据错误: {e}")
            logger.error(f"SQL错误详情: {str(e)}")
            # 返回空数据而不是错误
            return jsonify({
                'expense': {'labels': ['暂无数据'], 'datasets': [{'label': '支出', 'data': [0]}]},
                'income': {'labels': ['暂无数据'], 'datasets': [{'label': '收入', 'data': [0]}]}
            })
        finally:
            if connection and connection.is_connected():
                connection.close()
                
    except Exception as e:
        logger.error(f"获取类别统计图表数据异常: {str(e)}")
        import traceback
        logger.error(f"完整堆栈跟踪: {traceback.format_exc()}")
        # 返回空数据而不是错误
        return jsonify({
            'expense': {'labels': ['暂无数据'], 'datasets': [{'label': '支出', 'data': [0]}]},
            'income': {'labels': ['暂无数据'], 'datasets': [{'label': '收入', 'data': [0]}]}
        })

def get_simple_category_data(cursor, where_clause, params):
    """获取简单的类别统计数据（返回统一的数据结构）"""
    # 支出类别统计
    expense_query = f"""
        SELECT 
            category,
            COALESCE(SUM(amount), 0) as total_amount
        FROM daily_accounts 
        WHERE {where_clause} AND record_type = '支出'
        GROUP BY category
        ORDER BY total_amount DESC
        LIMIT 10
    """
    
    # 收入类别统计
    income_query = f"""
        SELECT 
            category,
            COALESCE(SUM(amount), 0) as total_amount
        FROM daily_accounts 
        WHERE {where_clause} AND record_type = '收入'
        GROUP BY category
        ORDER BY total_amount DESC
        LIMIT 10
    """
    
    cursor.execute(expense_query, params)
    expense_data = cursor.fetchall()
    
    cursor.execute(income_query, params)
    income_data = cursor.fetchall()
    
    # 处理支出数据
    expense_labels = []
    expense_amounts = []
    
    for item in expense_data:
        if item['category']:
            expense_labels.append(str(item['category']))
            expense_amounts.append(float(item['total_amount']))
    
    # 处理收入数据
    income_labels = []
    income_amounts = []
    
    for item in income_data:
        if item['category']:
            income_labels.append(str(item['category']))
            income_amounts.append(float(item['total_amount']))
    
    # 如果数据为空，提供默认值
    if not expense_labels or not expense_amounts:
        expense_labels = ['暂无数据']
        expense_amounts = [0]
    
    if not income_labels or not income_amounts:
        income_labels = ['暂无数据']
        income_amounts = [0]
    
    # 构建返回结果 - 统一使用datasets结构
    result = {
        'expense': {
            'labels': expense_labels,
            'datasets': [{
                'label': '支出金额',
                'data': expense_amounts
            }]
        },
        'income': {
            'labels': income_labels,
            'datasets': [{
                'label': '收入金额',
                'data': income_amounts
            }]
        }
    }
    
    logger.info(f"简单类别统计数据处理完成 - 支出类别数: {len(expense_labels)}, 收入类别数: {len(income_labels)}")
    return result

#添加一个辅助函数来确保数据格式统一（可选）：
def format_category_data(data):
    """格式化类别统计数据，确保返回统一的数据结构"""
    if not data:
        return {
            'expense': {
                'labels': ['暂无数据'],
                'datasets': [{'label': '支出金额', 'data': [0]}]
            },
            'income': {
                'labels': ['暂无数据'],
                'datasets': [{'label': '收入金额', 'data': [0]}]
            }
        }
    
    # 统一数据结构为datasets格式
    if 'expense' in data:
        if 'datasets' in data['expense']:
            # 已经是最新格式，确保每个dataset有label
            for dataset in data['expense']['datasets']:
                if 'label' not in dataset:
                    dataset['label'] = '支出金额'
        elif 'data' in data['expense'] and isinstance(data['expense']['data'], list):
            # 旧格式：转换为新格式
            data['expense']['datasets'] = [{
                'label': '支出金额',
                'data': data['expense']['data']
            }]
            # 删除旧的data字段
            if 'data' in data['expense']:
                del data['expense']['data']
    
    if 'income' in data:
        if 'datasets' in data['income']:
            # 已经是最新格式，确保每个dataset有label
            for dataset in data['income']['datasets']:
                if 'label' not in dataset:
                    dataset['label'] = '收入金额'
        elif 'data' in data['income'] and isinstance(data['income']['data'], list):
            # 旧格式：转换为新格式
            data['income']['datasets'] = [{
                'label': '收入金额',
                'data': data['income']['data']
            }]
            # 删除旧的data字段
            if 'data' in data['income']:
                del data['income']['data']
    
    return data
    
def get_time_based_category_data(cursor, where_clause, params, time_range):
    """获取基于时间维度的类别统计数据（修复版，按类别分组）"""
    try:
        # 确定时间分组字段
        if time_range == 'yearly':
            time_group = 'YEAR(account_date)'
            time_label = 'YEAR(account_date) as time_period'
            order_field = 'YEAR(account_date)'
        elif time_range == 'quarterly':
            time_group = 'YEAR(account_date), QUARTER(account_date)'
            time_label = 'CONCAT(YEAR(account_date), "年第", QUARTER(account_date), "季度") as time_period'
            order_field = 'YEAR(account_date), QUARTER(account_date)'
        elif time_range == 'monthly':
            time_group = 'YEAR(account_date), MONTH(account_date)'
            time_label = 'CONCAT(YEAR(account_date), "年", LPAD(MONTH(account_date), 2, "0"), "月") as time_period'
            order_field = 'YEAR(account_date), MONTH(account_date)'
        else:
            # 默认为年度
            time_group = 'YEAR(account_date)'
            time_label = 'YEAR(account_date) as time_period'
            order_field = 'YEAR(account_date)'
        
        # 获取所有时间段
        period_query = f"""
            SELECT DISTINCT {time_label}
            FROM daily_accounts 
            WHERE {where_clause}
            ORDER BY {order_field}
        """
        cursor.execute(period_query, params)
        periods_data = cursor.fetchall()
        periods = [str(item['time_period']) for item in periods_data if item['time_period']]
        
        # 如果没有时间段数据，返回空结果
        if not periods:
            return {
                'expense': {
                    'labels': [],
                    'datasets': []
                },
                'income': {
                    'labels': [],
                    'datasets': []
                }
            }
        
        # ==================== 修复：按类别查询支出数据 ====================
        expense_category_query = f"""
            SELECT 
                {time_label},
                category,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            WHERE {where_clause} AND record_type = '支出'
            GROUP BY {time_group}, category
            ORDER BY {order_field}, category
        """
        
        cursor.execute(expense_category_query, params)
        expense_data_raw = cursor.fetchall()
        
        # ==================== 修复：按类别查询收入数据 ====================
        income_category_query = f"""
            SELECT 
                {time_label},
                category,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            WHERE {where_clause} AND record_type = '收入'
            GROUP BY {time_group}, category
            ORDER BY {order_field}, category
        """
        
        cursor.execute(income_category_query, params)
        income_data_raw = cursor.fetchall()
        
        # 获取所有支出类别
        expense_categories = sorted(list(set([item['category'] for item in expense_data_raw if item['category']])))
        income_categories = sorted(list(set([item['category'] for item in income_data_raw if item['category']])))
        
        # 构建支出数据
        expense_datasets = []
        for category in expense_categories:
            category_data = [0] * len(periods)
            
            # 填充每个时间段的金额
            for item in expense_data_raw:
                if item['category'] == category:
                    period_idx = periods.index(str(item['time_period']))
                    category_data[period_idx] = float(item['total_amount'])
            
            expense_datasets.append({
                'label': category,
                'data': category_data
            })
        
        # 构建收入数据
        income_datasets = []
        for category in income_categories:
            category_data = [0] * len(periods)
            
            # 填充每个时间段的金额
            for item in income_data_raw:
                if item['category'] == category:
                    period_idx = periods.index(str(item['time_period']))
                    category_data[period_idx] = float(item['total_amount'])
            
            income_datasets.append({
                'label': category,
                'data': category_data
            })
        
        # 如果数据为空，提供默认值
        if not expense_datasets and expense_categories:
            # 如果有类别但没有数据，创建空数据集
            for category in expense_categories[:3]:  # 只显示前3个类别
                expense_datasets.append({
                    'label': category,
                    'data': [0] * len(periods)
                })
        
        if not income_datasets and income_categories:
            # 如果有类别但没有数据，创建空数据集
            for category in income_categories[:3]:  # 只显示前3个类别
                income_datasets.append({
                    'label': category,
                    'data': [0] * len(periods)
                })
        
        # 构建返回结果
        result = {
            'expense': {
                'labels': periods,
                'datasets': expense_datasets
            },
            'income': {
                'labels': periods,
                'datasets': income_datasets
            }
        }
        
        logger.info(f"时间维度类别统计数据处理完成 - 时间段数: {len(periods)}, 支出类别数: {len(expense_categories)}, 收入类别数: {len(income_categories)}")
        return result
        
    except Exception as e:
        logger.error(f"时间维度类别统计数据处理错误: {str(e)}")
        import traceback
        logger.error(f"完整堆栈跟踪: {traceback.format_exc()}")
        # 返回空数据而不是错误
        return {
            'expense': {
                'labels': ['暂无数据'],
                'datasets': [{'label': '无数据', 'data': [0]}]
            },
            'income': {
                'labels': ['暂无数据'],
                'datasets': [{'label': '无数据', 'data': [0]}]
            }
        }
    
# ===================== 新增：子类别金额统计API =====================
@app.route('/api/account/statistics/subcategory')
@login_required
def get_subcategory_statistics():
    """获取子类别金额统计"""
    try:
        subcategory = request.args.get('subcategory', '全部')
        owner = request.args.get('owner', '全部')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 验证必要参数
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': '开始日期和结束日期不能为空'})
        
        logger.info(f"子类别统计查询 - 子类别: {subcategory}, 所属人: {owner}, 日期: {start_date} 至 {end_date}")
        
        connection = create_connection()
        if not connection:
            return jsonify({'success': False, 'message': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = ["account_date BETWEEN %s AND %s"]
        params = [start_date, end_date]
        
        if subcategory != '全部':
            where_conditions.append("subcategory = %s")
            params.append(subcategory)
        
        if owner != '全部':
            where_conditions.append("owner = %s")
            params.append(owner)
        
        where_clause = " AND ".join(where_conditions)
        
        # 总统计查询
        total_query = f"""
            SELECT 
                COUNT(*) as record_count,
                COALESCE(SUM(amount), 0) as total_amount,
                COALESCE(SUM(CASE WHEN record_type = '支出' THEN amount ELSE 0 END), 0) as expense_amount,
                COALESCE(SUM(CASE WHEN record_type = '收入' THEN amount ELSE 0 END), 0) as income_amount
            FROM daily_accounts 
            WHERE {where_clause}
        """
        
        cursor.execute(total_query, params)
        total_stats = cursor.fetchone()
        
        # 计算净收入
        net_amount = float(total_stats['income_amount']) - float(total_stats['expense_amount'])
        
        # 获取详细记录（用于结果说明）
        detail_query = f"""
            SELECT record_type, category, subcategory, amount, account_date, owner
            FROM daily_accounts 
            WHERE {where_clause}
            ORDER BY account_date DESC
            LIMIT 100
        """
        
        cursor.execute(detail_query, params)
        detail_records = cursor.fetchall()
        
        cursor.close()
        
        result = {
            'success': True,
            'record_count': total_stats['record_count'],
            'total_amount': float(total_stats['total_amount']),
            'expense_amount': float(total_stats['expense_amount']),
            'income_amount': float(total_stats['income_amount']),
            'net_amount': net_amount,
            'detail_records': detail_records[:10]  # 只返回前10条记录用于参考
        }
        
        logger.info(f"子类别统计完成 - 记录数: {total_stats['record_count']}, 总金额: {total_stats['total_amount']}")
        
        return jsonify(result)
        
    except Error as e:
        logger.error(f"子类别统计查询错误: {e}")
        return jsonify({'success': False, 'message': f'统计查询失败: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"子类别统计异常: {str(e)}")
        return jsonify({'success': False, 'message': f'统计处理失败: {str(e)}'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/account/statistics/subcategory/export')
@login_required
def export_subcategory_statistics():
    """导出子类别统计结果"""
    try:
        subcategory = request.args.get('subcategory', '全部')
        owner = request.args.get('owner', '全部')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = ["account_date BETWEEN %s AND %s"]
        params = [start_date, end_date]
        
        if subcategory != '全部':
            where_conditions.append("subcategory = %s")
            params.append(subcategory)
        
        if owner != '全部':
            where_conditions.append("owner = %s")
            params.append(owner)
        
        where_clause = " AND ".join(where_conditions)
        
        # 查询详细记录
        detail_query = f"""
            SELECT record_type, category, subcategory, amount, account_date, description, payment_method, owner
            FROM daily_accounts 
            WHERE {where_clause}
            ORDER BY account_date DESC, record_type
        """
        
        cursor.execute(detail_query, params)
        records = cursor.fetchall()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "子类别统计"
        
        # 设置样式
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # 写入标题
        ws['A1'] = f"子类别金额统计报告 - {start_date} 至 {end_date}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # 写入筛选条件
        condition_text = f"筛选条件: 子类别={subcategory}, 所属人={owner}"
        ws['A2'] = condition_text
        ws.merge_cells('A2:H2')
        
        # 写入表头
        headers = ['记录类型', '类别', '子类别', '金额', '日期', '描述', '支付方式', '所属人']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row, record in enumerate(records, 5):
            ws.cell(row=row, column=1, value=record['record_type']).border = border
            ws.cell(row=row, column=2, value=record['category']).border = border
            ws.cell(row=row, column=3, value=record['subcategory'] or '').border = border
            ws.cell(row=row, column=4, value=float(record['amount'])).border = border
            ws.cell(row=row, column=5, value=record['account_date'].strftime('%Y-%m-%d') if not isinstance(record['account_date'], str) else record['account_date']).border = border
            ws.cell(row=row, column=6, value=record['description'] or '').border = border
            ws.cell(row=row, column=7, value=record['payment_method'] or '现金').border = border
            ws.cell(row=row, column=8, value=record['owner'] or '郭宁').border = border
        
        # 写入汇总信息
        summary_row = len(records) + 7
        ws.cell(row=summary_row, column=1, value="汇总统计").font = header_font
        ws.cell(row=summary_row + 1, column=1, value="总记录数")
        ws.cell(row=summary_row + 1, column=2, value=len(records))
        
        total_amount = sum(float(record['amount']) for record in records)
        expense_amount = sum(float(record['amount']) for record in records if record['record_type'] == '支出')
        income_amount = sum(float(record['amount']) for record in records if record['record_type'] == '收入')
        
        ws.cell(row=summary_row + 2, column=1, value="总金额")
        ws.cell(row=summary_row + 2, column=2, value=total_amount)
        
        ws.cell(row=summary_row + 3, column=1, value="支出总额")
        ws.cell(row=summary_row + 3, column=2, value=expense_amount)
        
        ws.cell(row=summary_row + 4, column=1, value="收入总额")
        ws.cell(row=summary_row + 4, column=2, value=income_amount)
        
        ws.cell(row=summary_row + 5, column=1, value="净收入")
        ws.cell(row=summary_row + 5, column=2, value=income_amount - expense_amount)
        
        # 设置列宽
        column_widths = [12, 15, 15, 12, 12, 25, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 设置金额格式
        for row in range(5, len(records) + 5):
            ws.cell(row=row, column=4).number_format = '0.00'
        for row in range(summary_row + 2, summary_row + 6):
            ws.cell(row=row, column=2).number_format = '0.00'
        
        cursor.close()
        
        # 创建内存文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"子类别统计_{start_date}_至_{end_date}.xlsx"
        if subcategory != '全部':
            filename = f"子类别统计_{subcategory}_{start_date}_至_{end_date}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"导出子类别统计错误: {str(e)}")
        return jsonify({'error': '导出统计结果失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 调试路由 =====================
@app.route('/api/debug/database_status')
@login_required
def debug_database_status():
    """调试数据库状态"""
    try:
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 检查表是否存在
        cursor.execute("SHOW TABLES LIKE 'daily_accounts'")
        accounts_table_exists = cursor.fetchone() is not None
        
        # 检查表结构
        table_info = {}
        if accounts_table_exists:
            cursor.execute("DESCRIBE daily_accounts")
            table_info['daily_accounts'] = cursor.fetchall()
            
            # 检查是否有数据
            cursor.execute("SELECT COUNT(*) as count FROM daily_accounts")
            table_info['daily_accounts_count'] = cursor.fetchone()['count']
        
        cursor.close()
        
        return jsonify({
            'accounts_table_exists': accounts_table_exists,
            'table_info': table_info
        })
        
    except Error as e:
        logger.error(f"调试数据库状态错误: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/debug/chart_data_verify')
@login_required
def debug_chart_data_verify():
    """调试图表数据验证"""
    try:
        start_date = request.args.get('start_date', '2024-01-01')
        end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # 查询郭宁的收入
        guoning_income_query = """
            SELECT COALESCE(SUM(amount), 0) as total 
            FROM daily_accounts 
            WHERE owner = '郭宁' AND record_type = '收入' 
            AND account_date BETWEEN %s AND %s
        """
        cursor.execute(guoning_income_query, (start_date, end_date))
        guoning_income = cursor.fetchone()['total']
        
        # 查询郭宁的支出
        guoning_expense_query = """
            SELECT COALESCE(SUM(amount), 0) as total 
            FROM daily_accounts 
            WHERE owner = '郭宁' AND record_type = '支出' 
            AND account_date BETWEEN %s AND %s
        """
        cursor.execute(guoning_expense_query, (start_date, end_date))
        guoning_expense = cursor.fetchone()['total']
        
        # 查询李佳慧的收入
        lijiahui_income_query = """
            SELECT COALESCE(SUM(amount), 0) as total 
            FROM daily_accounts 
            WHERE owner = '李佳慧' AND record_type = '收入' 
            AND account_date BETWEEN %s AND %s
        """
        cursor.execute(lijiahui_income_query, (start_date, end_date))
        lijiahui_income = cursor.fetchone()['total']
        
        # 查询李佳慧的支出
        lijiahui_expense_query = """
            SELECT COALESCE(SUM(amount), 0) as total 
            FROM daily_accounts 
            WHERE owner = '李佳慧' AND record_type = '支出' 
            AND account_date BETWEEN %s AND %s
        """
        cursor.execute(lijiahui_expense_query, (start_date, end_date))
        lijiahui_expense = cursor.fetchone()['total']
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': {
                '郭宁': {
                    '收入': float(guoning_income),
                    '支出': float(guoning_expense)
                },
                '李佳慧': {
                    '收入': float(lijiahui_income),
                    '支出': float(lijiahui_expense)
                }
            }
        })
        
    except Exception as e:
        logger.error(f"调试图表数据验证错误: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/account/export/test')
@login_required
def test_export_api():
    """测试导出API是否正常工作"""
    try:
        return jsonify({
            'success': True,
            'message': '导出API正常工作',
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'supported_params': ['record_type', 'category', 'subcategory', 'start_date', 'end_date', 'owner']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 调试路由
@app.route('/api/debug/reset_and_verify', methods=['POST'])
def debug_reset_and_verify():
    """调试用的密码重置和验证路由"""
    data = request.json
    username = data.get('username', 'admin')
    new_password = data.get('new_password', '123456789')
    
    logger.info("=== 详细调试密码重置开始 ===")
    
    # 使用修复的版本
    result = reset_user_password(username, new_password)
    
    # 立即验证
    verification_result = verify_login(username, new_password)
    
    return jsonify({
        'reset_success': result,
        'login_verification': verification_result,
        'message': '请查看应用日志了解详细过程'
    })

@app.route('/api/debug/verify_password', methods=['POST'])
def debug_verify_password():
    """手动验证密码哈希（仅用于调试）"""
    data = request.json
    username = data.get('username', 'admin')
    password = data.get('password', '')
    
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'})
    
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT password_hash, password_salt 
            FROM user_security 
            WHERE username = %s
        """, (username,))
        
        result = cursor.fetchone()
        
        if not result:
            return jsonify({'error': '用户不存在'})
        
        # 手动计算哈希
        calculated_hash = hash_password(password, result['password_salt'])
        
        return jsonify({
            'stored_hash': result['password_hash'],
            'stored_salt': result['password_salt'],
            'calculated_hash': calculated_hash,
            'match': result['password_hash'] == calculated_hash,
            'debug_info': {
                'password': password,
                'password_length': len(password),
                'salt_length': len(result['password_salt'])
            }
        })
        
    except Error as e:
        logger.error(f"调试验证错误: {e}")
        return jsonify({'error': str(e)})
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：所属人管理功能 =====================

@app.route('/api/owners', methods=['GET'])
@login_required
def get_owners():
    """获取所有所属人"""
    connection = create_connection()
    if not connection:
        return jsonify({'error': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        owners_set = set(['郭宁', '李佳慧'])  # 默认所属人
        
        try:
            # 从记账记录中获取所有不重复的所属人
            cursor.execute("""
                SELECT DISTINCT owner 
                FROM daily_accounts 
                WHERE owner IS NOT NULL AND owner != ''
                ORDER BY owner
            """)
            
            for row in cursor.fetchall():
                if row['owner']:
                    owners_set.add(row['owner'])
        except Exception as e:
            logger.warning(f"从daily_accounts获取所属人失败: {e}")
        
        try:
            # 从礼尚往来记录中获取所有不重复的所属人
            cursor.execute("""
                SELECT DISTINCT owner 
                FROM gift_records 
                WHERE owner IS NOT NULL AND owner != ''
                ORDER BY owner
            """)
            
            for row in cursor.fetchall():
                if row['owner']:
                    owners_set.add(row['owner'])
        except Exception as e:
            logger.warning(f"从gift_records获取所属人失败: {e}")
        
        # 获取之前保存的所属人列表
        try:
            cursor.execute("""
                SELECT config_value 
                FROM system_config 
                WHERE config_key = 'account_owners'
            """)
            stored_owners = cursor.fetchone()
            
            # 如果数据库中有保存的所属人列表，合并
            if stored_owners and stored_owners['config_value']:
                try:
                    stored_list = json.loads(stored_owners['config_value'])
                    owners_set.update(stored_list)
                except json.JSONDecodeError:
                    pass
        except Exception as e:
            logger.warning(f"从system_config获取所属人失败: {e}")
        
        owners = sorted(list(owners_set))
        cursor.close()
        
        return jsonify({'owners': owners})
        
    except Error as e:
        logger.error(f"获取所属人列表错误: {e}")
        return jsonify({'error': '获取所属人列表失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/owners', methods=['POST'])
@login_required
def add_owner():
    """添加新的所属人"""
    data = request.json
    new_owner = data.get('owner', '').strip()
    
    if not new_owner:
        return jsonify({'success': False, 'message': '所属人名称不能为空'})
    
    connection = create_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 获取现有的所属人列表
        cursor.execute("""
            SELECT config_value 
            FROM system_config 
            WHERE config_key = 'account_owners'
        """)
        stored_owners = cursor.fetchone()
        
        owners_list = []
        if stored_owners and stored_owners['config_value']:
            try:
                owners_list = json.loads(stored_owners['config_value'])
            except json.JSONDecodeError:
                owners_list = []
        
        # 检查是否已存在
        if new_owner in owners_list:
            return jsonify({'success': False, 'message': '所属人已存在'})
        
        # 添加新的所属人
        owners_list.append(new_owner)
        owners_list.sort()
        
        # 保存到数据库
        cursor.execute("""
            INSERT INTO system_config (config_key, config_value) 
            VALUES ('account_owners', %s)
            ON DUPLICATE KEY UPDATE config_value = %s, updated_at = CURRENT_TIMESTAMP
        """, (json.dumps(owners_list, ensure_ascii=False), json.dumps(owners_list, ensure_ascii=False)))
        
        connection.commit()
        cursor.close()
        
        # 记录操作日志
        log_operation("SYSTEM", f"添加所属人: {new_owner}", user_name=session.get('username', 'admin'))
        
        return jsonify({'success': True, 'message': '所属人添加成功'})
        
    except Error as e:
        logger.error(f"添加所属人错误: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': '添加所属人失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/owners/<owner_name>', methods=['DELETE'])
@login_required
def delete_owner(owner_name):
    """删除所属人"""
    # 防止删除默认所属人
    if owner_name in ['郭宁', '李佳慧']:
        return jsonify({'success': False, 'message': '不能删除默认所属人'})
    
    connection = create_connection()
    if not connection:
        return jsonify({'success': False, 'message': '数据库连接失败'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        
        # 获取现有的所属人列表
        cursor.execute("""
            SELECT config_value 
            FROM system_config 
            WHERE config_key = 'account_owners'
        """)
        stored_owners = cursor.fetchone()
        
        owners_list = []
        if stored_owners and stored_owners['config_value']:
            try:
                owners_list = json.loads(stored_owners['config_value'])
            except json.JSONDecodeError:
                owners_list = []
        
        # 检查是否存在于列表中
        if owner_name not in owners_list:
            return jsonify({'success': False, 'message': '所属人不存在'})
        
        # 从列表中移除
        owners_list.remove(owner_name)
        
        # 保存到数据库
        cursor.execute("""
            INSERT INTO system_config (config_key, config_value) 
            VALUES ('account_owners', %s)
            ON DUPLICATE KEY UPDATE config_value = %s, updated_at = CURRENT_TIMESTAMP
        """, (json.dumps(owners_list, ensure_ascii=False), json.dumps(owners_list, ensure_ascii=False)))
        
        connection.commit()
        cursor.close()
        
        # 记录操作日志
        log_operation("SYSTEM", f"删除所属人: {owner_name}", user_name=session.get('username', 'admin'))
        
        return jsonify({'success': True, 'message': '所属人删除成功'})
        
    except Error as e:
        logger.error(f"删除所属人错误: {e}")
        connection.rollback()
        return jsonify({'success': False, 'message': '删除所属人失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：按所属人进行深度统计 =====================

@app.route('/api/account/statistics/by_owner')
@login_required
def get_statistics_by_owner():
    """按所属人进行深度统计"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        stat_type = request.args.get('type', 'monthly')  # monthly, quarterly, yearly
        
        # 如果没有提供日期范围，默认使用最近一年
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 构建查询条件
        where_conditions = ["account_date BETWEEN %s AND %s"]
        params = [start_date, end_date]
        
        where_clause = " AND ".join(where_conditions)
        
        # 根据统计类型构建查询
        if stat_type == 'monthly':
            group_by = "YEAR(account_date), MONTH(account_date), owner, record_type"
            period_select = "CONCAT(YEAR(account_date), '年', LPAD(MONTH(account_date), 2, '0'), '月') as period"
            period_order = "YEAR(account_date), MONTH(account_date)"
        elif stat_type == 'quarterly':
            group_by = "YEAR(account_date), QUARTER(account_date), owner, record_type"
            period_select = "CONCAT(YEAR(account_date), '年', QUARTER(account_date), '季度') as period"
            period_order = "YEAR(account_date), QUARTER(account_date)"
        elif stat_type == 'yearly':
            group_by = "YEAR(account_date), owner, record_type"
            period_select = "CONCAT(YEAR(account_date), '年') as period"
            period_order = "YEAR(account_date)"
        else:
            return jsonify({'error': '不支持的统计类型'}), 400
        
        # 查询数据
        query = f"""
            SELECT 
                {period_select},
                owner,
                record_type,
                COUNT(*) as count,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            WHERE {where_clause}
            GROUP BY {group_by}
            ORDER BY {period_order}, owner, record_type
        """
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # 获取所有所属人
        owners_query = "SELECT DISTINCT owner FROM daily_accounts WHERE owner IS NOT NULL AND owner != ''"
        cursor.execute(owners_query)
        owners_result = cursor.fetchall()
        all_owners = [owner['owner'] for owner in owners_result]
        
        # 获取所有时间段
        periods_query = f"""
            SELECT DISTINCT {period_select}
            FROM daily_accounts 
            WHERE {where_clause}
            ORDER BY {period_order}
        """
        cursor.execute(periods_query, params)
        periods_result = cursor.fetchall()
        periods = [period['period'] for period in periods_result]
        
        # 处理数据，按所属人和时间段组织
        statistics = {}
        for period in periods:
            statistics[period] = {}
            for owner in all_owners:
                statistics[period][owner] = {
                    '支出': {'count': 0, 'amount': 0.0},
                    '收入': {'count': 0, 'amount': 0.0},
                    '净收入': 0.0
                }
        
        # 填充数据
        for row in results:
            period = row['period']
            owner = row['owner']
            record_type = row['record_type']
            
            if period in statistics and owner in statistics[period]:
                statistics[period][owner][record_type]['count'] = row['count']
                statistics[period][owner][record_type]['amount'] = float(row['total_amount'])
        
        # 计算净收入
        for period in statistics:
            for owner in statistics[period]:
                income = statistics[period][owner]['收入']['amount']
                expense = statistics[period][owner]['支出']['amount']
                statistics[period][owner]['净收入'] = income - expense
        
        # 计算总计
        totals = {}
        for owner in all_owners:
            totals[owner] = {
                '支出': {'count': 0, 'amount': 0.0},
                '收入': {'count': 0, 'amount': 0.0},
                '净收入': 0.0
            }
        
        # 汇总所有时间段的数据
        for period in statistics:
            for owner in statistics[period]:
                for record_type in ['支出', '收入']:
                    totals[owner][record_type]['count'] += statistics[period][owner][record_type]['count']
                    totals[owner][record_type]['amount'] += statistics[period][owner][record_type]['amount']
        
        # 计算总计的净收入
        for owner in totals:
            totals[owner]['净收入'] = totals[owner]['收入']['amount'] - totals[owner]['支出']['amount']
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'periods': periods,
            'owners': all_owners,
            'statistics': statistics,
            'totals': totals,
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'type': stat_type
            }
        })
        
    except Error as e:
        logger.error(f"按所属人统计错误: {e}")
        return jsonify({'success': False, 'message': '统计失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

# ===================== 新增：所属人对比统计 =====================

@app.route('/api/account/statistics/owner_comparison')
@login_required
def get_owner_comparison_statistics():
    """所属人对比统计"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 如果没有提供日期范围，默认使用最近一年
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        connection = create_connection()
        if not connection:
            return jsonify({'error': '数据库连接失败'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # 查询按类别统计的所属人数据
        category_query = """
            SELECT 
                category,
                owner,
                record_type,
                COUNT(*) as count,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            WHERE account_date BETWEEN %s AND %s
            GROUP BY category, owner, record_type
            ORDER BY category, owner, record_type
        """
        
        cursor.execute(category_query, (start_date, end_date))
        category_results = cursor.fetchall()
        
        # 查询按月统计的所属人数据
        monthly_query = """
            SELECT 
                DATE_FORMAT(account_date, '%Y-%m') as month,
                owner,
                record_type,
                COUNT(*) as count,
                COALESCE(SUM(amount), 0) as total_amount
            FROM daily_accounts 
            WHERE account_date BETWEEN %s AND %s
            GROUP BY DATE_FORMAT(account_date, '%Y-%m'), owner, record_type
            ORDER BY month, owner, record_type
        """
        
        cursor.execute(monthly_query, (start_date, end_date))
        monthly_results = cursor.fetchall()
        
        # 获取所有所属人
        owners_query = "SELECT DISTINCT owner FROM daily_accounts WHERE owner IS NOT NULL AND owner != ''"
        cursor.execute(owners_query)
        owners_result = cursor.fetchall()
        all_owners = [owner['owner'] for owner in owners_result]
        
        # 处理按类别统计的数据
        category_stats = {}
        for row in category_results:
            category = row['category']
            owner = row['owner']
            record_type = row['record_type']
            
            if category not in category_stats:
                category_stats[category] = {}
            
            if owner not in category_stats[category]:
                category_stats[category][owner] = {
                    '支出': 0.0,
                    '收入': 0.0
                }
            
            category_stats[category][owner][record_type] = float(row['total_amount'])
        
        # 处理按月统计的数据
        monthly_stats = {}
        for row in monthly_results:
            month = row['month']
            owner = row['owner']
            record_type = row['record_type']
            
            if month not in monthly_stats:
                monthly_stats[month] = {}
            
            if owner not in monthly_stats[month]:
                monthly_stats[month][owner] = {
                    '支出': 0.0,
                    '收入': 0.0,
                    '净收入': 0.0
                }
            
            monthly_stats[month][owner][record_type] = float(row['total_amount'])
        
        # 计算每月的净收入
        for month in monthly_stats:
            for owner in monthly_stats[month]:
                income = monthly_stats[month][owner]['收入']
                expense = monthly_stats[month][owner]['支出']
                monthly_stats[month][owner]['净收入'] = income - expense
        
        # 计算总计
        totals = {}
        for owner in all_owners:
            totals[owner] = {
                '支出': 0.0,
                '收入': 0.0,
                '净收入': 0.0,
                '记录数': 0,
                '支出占比': 0.0,
                '收入占比': 0.0
            }
        
        # 汇总数据
        for row in monthly_results:
            owner = row['owner']
            record_type = row['record_type']
            totals[owner][record_type] += float(row['total_amount'])
            totals[owner]['记录数'] += row['count']
        
        # 计算净收入和占比
        total_expense = sum([totals[owner]['支出'] for owner in totals])
        total_income = sum([totals[owner]['收入'] for owner in totals])
        
        for owner in totals:
            totals[owner]['净收入'] = totals[owner]['收入'] - totals[owner]['支出']
            if total_expense > 0:
                totals[owner]['支出占比'] = (totals[owner]['支出'] / total_expense * 100)
            if total_income > 0:
                totals[owner]['收入占比'] = (totals[owner]['收入'] / total_income * 100)
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'category_stats': category_stats,
            'monthly_stats': monthly_stats,
            'totals': totals,
            'owners': all_owners,
            'filters': {
                'start_date': start_date,
                'end_date': end_date
            }
        })
        
    except Error as e:
        logger.error(f"所属人对比统计错误: {e}")
        return jsonify({'success': False, 'message': '统计失败'}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()



# 数据库连接池
db_pool = None

# 优化数据库连接池配置
def init_database_pool():
    global db_pool
    try:
        db_pool = pooling.MySQLConnectionPool(
            pool_name="gift_pool",
            pool_size=10,  # 增加连接数
            pool_reset_session=True,
            **DB_CONFIG
        )
        logger.info("数据库连接池初始化成功")
    except Error as e:
        logger.error(f"数据库连接池初始化失败: {e}")
        
def safe_execute(cursor, query, params=None, timeout=30):
    """安全执行SQL查询（带超时）"""
    try:
        # 设置查询超时
        cursor.execute("SET SESSION MAX_EXECUTION_TIME=%s", (timeout * 1000,))
        cursor.execute(query, params or ())
        if cursor.with_rows:
            result = cursor.fetchall()
            return result
        return None
    except Error as e:
        raise e

def create_connection():
    """创建数据库连接（使用连接池）"""
    try:
        if db_pool:
            return db_pool.get_connection()
        else:
            return mysql.connector.connect(**DB_CONFIG)
    except Error as e:
        logger.error(f"数据库连接失败: {str(e)}")
        return None


if __name__ == '__main__':
    # 初始化数据库
    if init_database():
        init_database_pool()  # 初始化连接池
        logger.info("数据库初始化成功")
        logger.info("家庭记账管理系统已启动!")
        logger.info("访问地址: http://localhost:5000")
        logger.info("主页面: 记账管理")
        logger.info("二级页面: 礼尚往来记录管理 (/gift_management)")
        logger.info("✅ 密码使用安全的哈希方式存储")
        if not HAS_PINYIN:
            logger.info("提示：如需准确的姓名拼音排序，请安装pypinyin库")
            logger.info("运行: pip install pypinyin")
        
        try:
            # 启动Flask应用
            app.run(host='0.0.0.0', port=5000, debug=False)
        except Exception as e:
            logger.error(f"启动失败: {str(e)}")
    else:
        logger.error("数据库初始化失败，程序退出")